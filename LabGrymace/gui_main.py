import os
import re
import shutil
import warnings
import wx
import wx.lib.agw.hyperlink as hl
import wx.lib.scrolledpanel as scrolled
import pandas as pd
import numpy as np
import matplotlib
matplotlib.use('Agg')
import matplotlib.pyplot as plt
from pathlib import Path
from LabGrymace import __version__

# ============================================================================
# Calibration — DERIVED from the Figure 4 outputs (single source of truth; NOT hand-typed numbers).
#   weights W_*      = |slope| of each region's intensity_area dose-response   (Figure 4B-D 'Fit')
#   Z-ref mean/std   = Figure 4E 'Z_reference' sheet
#   0/100 anchors    = Figure 4E 'Anchors_stats' sheet  (BASELINE_RAW -> 0, CNO_1MG_RAW -> 100)
# The values are read from those Excel files, so there is no literal to hand-edit and the GUI stays
# consistent with the figures. By default it reads the two Excel files bundled inside this package
# (LabGrymace/calibration/), so the install is self-contained. To track a live/regenerated Figure 4
# instead, set the environment variable LABGRYMACE_FIGURE4_DIR to that directory (it must contain
# Figure4B-D/figure4bd_data.xlsx and Figure4E/dose_response_data.xlsx); the GUI then follows it.
# ============================================================================
_FIG4_BUNDLE = Path(__file__).resolve().parent / 'calibration'
_FIG4_LIVE   = os.environ.get('LABGRYMACE_FIGURE4_DIR')   # optional override -> a live Figure 4 dir

def _read_calibration(bd_xlsx, de_xlsx):
    '''weights from Figure 4B-D slopes; mean/std + anchors from Figure 4E (all derived, no literals).'''
    fit = pd.read_excel(bd_xlsx, sheet_name='Fit').set_index('Behavior')['slope']
    W = (abs(float(fit['Ear Resting State'])), abs(float(fit['Orbital Tightening'])),
         abs(float(fit['Nose Bulging'])))
    z = pd.read_excel(de_xlsx, sheet_name='Z_reference').set_index('Region')
    MS = tuple(float(z.loc[r, c]) for r in ('ear', 'eye', 'nose') for c in ('CNO_mean', 'CNO_std'))
    a = pd.read_excel(de_xlsx, sheet_name='Anchors_stats').set_index('Item')['Value']
    r0   = float(next(v for k, v in a.items() if str(k).startswith('BASELINE_RAW')))
    r100 = float(next(v for k, v in a.items() if str(k).startswith('CNO_1MG_RAW')))
    return W, MS, (r0, r100)

def _load_calibration():
    sources = []                                   # an explicit live Figure 4 dir wins; then the bundle
    if _FIG4_LIVE:
        live = Path(_FIG4_LIVE)
        sources.append((live / 'Figure4B-D' / 'figure4bd_data.xlsx',
                        live / 'Figure4E' / 'dose_response_data.xlsx'))
    sources.append((_FIG4_BUNDLE / 'figure4bd_data.xlsx', _FIG4_BUNDLE / 'dose_response_data.xlsx'))
    for bd, de in sources:
        try:
            if bd.exists() and de.exists():
                return _read_calibration(bd, de) + (str(bd.parent),)
        except Exception as e:
            warnings.warn(f'[calibration] failed to read from {bd.parent}: {e}')
    raise RuntimeError('No Figure 4 calibration source found (bundled copies missing?).')

(W_EAR, W_EYE, W_NOSE), \
(CNO_EAR_MEAN, CNO_EAR_STD, CNO_EYE_MEAN, CNO_EYE_STD, CNO_NOSE_MEAN, CNO_NOSE_STD), \
(BASELINE_RAW_SCORE, CNO_1MG_RAW_SCORE), CALIBRATION_SOURCE = _load_calibration()

# Temporal window size (frames per segment)
FRAME_WINDOW = 3000

# Event filter fallback threshold: if any channel has < this % of filtered
# rows vs total rows, fall back to all-events averaging for that animal.
FALLBACK_THRESHOLD_PCT = 3.0
USE_EVENT_FILTER = False   # False = all_events + mirror filter only; True = BL/OT/Bul masking

# Mirror-reflection filter is OPT-IN. Default OFF: load_raw_data does NOT remove
# any frames unless the caller (or GUI checkbox) explicitly turns it on. Turn it on
# only when the LabGym tracker may have jumped to mirror reflections. See the GUI
# "Apply mirror-reflection filter" checkbox and load_raw_data(apply_mirror_filter=...).
# NOTE: the Figure 4E / Figure 6 repro scripts call load_raw_data(..., apply_mirror_filter=True)
# explicitly, so the published figures are unaffected by this default.
APPLY_MIRROR_FILTER = False

# ── Adaptive nose mirror-reflection filter (4-condition) ─────────────────────
# Glass-tube and similar reflective setups cause the nose centroid to alternate
# between the real nose and its mirror image.  A frame is only removed when
# ALL FOUR conditions are satisfied simultaneously — this is deliberately
# conservative so that no legitimate nose movement is discarded.
#
# Condition 1 — high velocity:      velocity      > VEL_THRESH
# Condition 2 — near mirror point:  dist(pos, mirror_point) < SPATIAL_RADIUS
#   The mirror point is auto-detected as the densest spatial cluster among
#   all high-velocity nose frames for that animal.  If the cluster is not
#   strong enough (< SPATIAL_MIN_PCT of high-vel frames), condition 2 falls
#   back to True for all high-vel frames (no spatial data available).
# Condition 3 — large displacement: magnitude_area > MAG_THRESH
# Condition 4 — temporal recurrence: ≥ TEMPORAL_MIN_N other high-vel frames
#   exist within ±TEMPORAL_WINDOW frames  (rules out isolated fast movements)
#
# Pre-check (bimodal gate): the 4-condition filter only runs at all when
# > VEL_BIMODAL_PCT of NoseBul frames have velocity above VEL_THRESH.
# Normal animals (2–8 %) are skipped entirely; artifact animals (30–80 %)
# proceed to the per-frame 4-condition test.
#
# Position data (NoseX / NoseY) must be present in nose_summary.xlsx.
# Run "Generate Summary Files" with this version of LabGrymace to populate
# those columns from nose_all_centers.xlsx in the raw LabGym folder.

VEL_THRESH        = 10.0   # px/frame — high-velocity threshold (condition 1)
MAG_THRESH        = 0.5    # magnitude_area — large-displacement gate (condition 3)
VEL_BIMODAL_PCT   = 0.15   # fraction of NoseBul frames above VEL_THRESH needed
                            # to engage the filter at all
SPATIAL_RADIUS    = 50.0   # pixels — radius around mirror point (condition 2)
SPATIAL_MIN_PCT   = 0.40   # min fraction of high-vel frames that must cluster
                            # before the mirror point is considered reliable
TEMPORAL_WINDOW   = 30     # ±frames — neighbourhood for temporal check (cond 4)
TEMPORAL_MIN_N    = 2      # min high-vel neighbours required (condition 4)


# ============================================================================
# Data helpers
# ============================================================================

# Summary files produced by the Generate step
_SUMMARY_FILES = ('ear_summary.xlsx', 'eye_summary.xlsx', 'nose_summary.xlsx')
# Raw LabGym behavior subfolders (used only by the Generate step)
_RAW_MARKERS = ('fEarBL', 'sEarBL', 'fEyeOT', 'noseBul')


def _has_summary_files(path):
    '''True if all three summary files are present in path.'''
    return all((path / f).exists() for f in _SUMMARY_FILES)


def _is_raw_labgym_folder(path):
    '''True if path looks like a LabGym _processed output folder (has behavior subfolders).'''
    return any((path / b).exists() for b in _RAW_MARKERS)


# ── Pain Score window: only accept folders that already have summary files ──

def find_animals_in_folder(root_folder):
    '''
    Scan root_folder for animal subfolders that contain pre-generated summary files
    (ear_summary.xlsx, eye_summary.xlsx, nose_summary.xlsx).

    If root_folder itself has summary files, treat it as a single animal.

    Returns: [(default_name, folder_path_str), ...]
    '''
    root = Path(root_folder)

    if _has_summary_files(root):
        return [(root.name, str(root))]

    animals = []
    for entry in sorted(root.iterdir()):
        if entry.is_dir() and _has_summary_files(entry):
            animals.append((entry.name, str(entry)))
    return animals


# ── Generate step: find raw LabGym folders ──

def find_raw_folders(root_folder):
    '''
    Scan root_folder for raw LabGym _processed subfolders (contain fEarBL etc.).
    If root_folder itself is a raw folder, treat it as one animal.

    Returns: [(default_name, folder_path_str), ...]
    '''
    root = Path(root_folder)

    if _is_raw_labgym_folder(root):
        return [(root.name, str(root))]

    animals = []
    for entry in sorted(root.iterdir()):
        if entry.is_dir() and _is_raw_labgym_folder(entry):
            animals.append((entry.name, str(entry)))
    return animals


def generate_summary_files(folder_path, output_dir=None):
    '''
    Run FacialDataIntegrator + LoadDataAdvanced on a raw LabGym _processed folder
    to generate ear_summary.xlsx, eye_summary.xlsx, nose_summary.xlsx.

    output_dir: where to write the three summary files.
                Defaults to folder_path itself (alongside the raw data).
    Skips if all three already exist in output_dir.
    '''
    import tempfile
    from .loaddata import FacialDataIntegrator, LoadDataAdvanced

    folder = Path(folder_path)
    out = Path(output_dir) if output_dir else folder
    out.mkdir(parents=True, exist_ok=True)

    if all((out / f).exists() for f in _SUMMARY_FILES):
        return  # already done

    dataset_key = '_current'
    # Use a true local temp directory to avoid cloud-sync timeouts
    _tmpdir = tempfile.mkdtemp(prefix='labgrymace_')
    temp_root = Path(_tmpdir)

    # --- Step 1: FacialDataIntegrator ---
    integrator = FacialDataIntegrator.__new__(FacialDataIntegrator)
    integrator.base_path = str(folder)
    integrator.output_base = str(temp_root)
    integrator.datasets = {dataset_key: str(folder)}
    integrator.eye_behaviors = ['fEyeBL', 'fEyeOT', 'sEyeBL', 'sEyeOT']
    integrator.ear_behaviors = ['fEarBL', 'fEarPB', 'sEarBL', 'sEarPB']
    integrator.nose_behaviors = ['noseBL', 'noseBul']
    integrator.eye_parameters = [
        'eyes_acceleration.xlsx', 'eyes_intensity_area.xlsx', 'eyes_intensity_length.xlsx',
        'eyes_magnitude_area.xlsx', 'eyes_magnitude_length.xlsx', 'eyes_probability.xlsx',
        'eyes_speed.xlsx', 'eyes_velocity.xlsx', 'eyes_vigor_area.xlsx', 'eyes_vigor_length.xlsx',
    ]
    integrator.ear_parameters = [
        'ears_acceleration.xlsx', 'ears_intensity_area.xlsx', 'ears_intensity_length.xlsx',
        'ears_magnitude_area.xlsx', 'ears_magnitude_length.xlsx', 'ears_probability.xlsx',
        'ears_speed.xlsx', 'ears_velocity.xlsx', 'ears_vigor_area.xlsx', 'ears_vigor_length.xlsx',
    ]
    integrator.nose_parameters = [
        'nose_acceleration.xlsx', 'nose_intensity_area.xlsx', 'nose_intensity_length.xlsx',
        'nose_magnitude_area.xlsx', 'nose_magnitude_length.xlsx', 'nose_probability.xlsx',
        'nose_speed.xlsx', 'nose_velocity.xlsx', 'nose_vigor_area.xlsx', 'nose_vigor_length.xlsx',
    ]

    for feature in ['eyes', 'ears', 'nose']:
        data = integrator.integrate_feature_data(dataset_key, feature)
        if data and any(d for d in data.values()):
            integrator.create_integrated_excel(data, dataset_key, feature)

    # --- Step 2: LoadDataAdvanced ---
    lda = LoadDataAdvanced.__new__(LoadDataAdvanced)
    lda.base_path = str(folder)
    lda.summary_root = str(temp_root)
    lda.param_keys = [
        'intensity_area', 'intensity_length', 'magnitude_area', 'magnitude_length',
        'probability', 'speed', 'velocity', 'vigor_area', 'vigor_length',
    ]

    lda.build_ear_summary(dataset_key, str(folder))
    lda.build_eye_summary(dataset_key, str(folder))
    lda.build_nose_summary(dataset_key, str(folder))
    lda.enforce_sheet_labels(dataset_key)

    # Move summary files from temp dir to output_dir
    temp_dir = temp_root / dataset_key
    for fname in _SUMMARY_FILES:
        src = temp_dir / fname
        dst = out / fname
        if src.exists():
            shutil.move(str(src), str(dst))

    # Merge centroid positions into summary files (for 4-condition spatial filter)
    _merge_nose_positions(folder, out / 'nose_summary.xlsx')
    _merge_ear_positions(folder,  out / 'ear_summary.xlsx')
    _merge_eye_positions(folder,  out / 'eye_summary.xlsx')

    # Rename the view columns to a self-explanatory name and spell out their values,
    # so the tables read clearly without a key ("Pos"/"f"/"s" are opaque to a reader).
    # loaddata.py keeps its internal "Pos" naming; only the written output is renamed.
    try:
        _rename_pos_to_view(out)
    except Exception:
        pass

    # Add the YELLOW "merged intensity (feeds pain score)" column for reviewer
    # traceability. Mirror filter ON = the published Figure 4E / Figure 6 convention.
    try:
        add_pain_score_merged_column(out, apply_mirror_filter=True)
    except Exception:
        pass

    # Reviewer subsheets (PanelB_behavior / PanelC_painscore) so the generated files
    # carry the same traceability layout as the published summary tables.
    try:
        add_behavior_and_painscore_panels(out, apply_mirror_filter=True)
    except Exception:
        pass

    # Clean up local temp directory
    try:
        shutil.rmtree(str(_tmpdir))
    except Exception:
        pass


_POS_TO_VIEW = {'Ear0Pos': 'Ear0View', 'Ear1Pos': 'Ear1View',
                'Eye0Pos': 'Eye0View', 'Eye1Pos': 'Eye1View',
                'NosePos': 'NoseView'}
_VIEW_VALUES = {'f': 'frontal', 's': 'lateral'}


def _rename_pos_to_view(out_dir):
    '''Rename the "<region>Pos" columns of the generated summary files to
    "<region>View" and spell out their values (f -> frontal, s -> lateral).

    Only the written summary files are touched; loaddata.py keeps its internal
    "Pos" naming (it maps LabGym's raw headers). Nothing downstream reads these
    columns -- the analyses read the Event and parameter columns -- so this is a
    presentation-only change. Idempotent.'''
    import openpyxl
    out_dir = Path(out_dir)
    for fname in _SUMMARY_FILES:
        p = out_dir / fname
        if not p.exists():
            continue
        wb = openpyxl.load_workbook(str(p))
        touched = False
        for ws in wb.worksheets:
            for c in range(1, ws.max_column + 1):
                cell = ws.cell(row=1, column=c)
                if cell.value in _POS_TO_VIEW:
                    cell.value = _POS_TO_VIEW[cell.value]
                    touched = True
            if ws.title == 'Summary':
                hdr = [ws.cell(row=1, column=c).value for c in range(1, ws.max_column + 1)]
                for c, h in enumerate(hdr, 1):
                    if h and str(h).endswith('View'):
                        for r in range(2, ws.max_row + 1):
                            cell = ws.cell(row=r, column=c)
                            if cell.value in _VIEW_VALUES:
                                cell.value = _VIEW_VALUES[cell.value]
                                touched = True
        if touched:
            tmp = str(p) + '.writing.xlsx'
            wb.save(tmp)
            os.replace(tmp, str(p))
        wb.close()


def _merge_nose_positions(raw_folder, nose_summary_path):
    '''Read nose_all_centers.xlsx from the raw LabGym folder and write
    NoseX / NoseY columns into nose_summary.xlsx.

    nose_all_centers.xlsx has rows like:
        time/ID  |  0
        0.03     |  (736, 748)

    Alignment is done by matching row counts (both files have the same number
    of time-steps); if counts differ by exactly 1 the shorter is padded with
    NaN at the end.
    '''
    centers_path = Path(raw_folder) / 'nose_all_centers.xlsx'
    if not centers_path.exists() or not nose_summary_path.exists():
        return
    try:
        centers = pd.read_excel(centers_path)
        pos_col = centers.iloc[:, 1]   # column "0" — "(x, y)" strings or NaN

        def _parse(val):
            if pd.isna(val):
                return np.nan, np.nan
            s = str(val).strip().lstrip('(').rstrip(')')
            parts = s.split(',')
            if len(parts) != 2:
                return np.nan, np.nan
            try:
                return float(parts[0].strip()), float(parts[1].strip())
            except ValueError:
                return np.nan, np.nan

        parsed = [_parse(v) for v in pos_col]
        xs = [p[0] for p in parsed]
        ys = [p[1] for p in parsed]

        summary = pd.read_excel(nose_summary_path, sheet_name='Summary')
        n_sum = len(summary)
        n_cen = len(xs)

        # Align lengths (pad shorter with NaN)
        if n_cen < n_sum:
            xs += [np.nan] * (n_sum - n_cen)
            ys += [np.nan] * (n_sum - n_cen)
        summary['NoseX'] = xs[:n_sum]
        summary['NoseY'] = ys[:n_sum]

        with pd.ExcelWriter(str(nose_summary_path), engine='openpyxl') as w:
            summary.to_excel(w, sheet_name='Summary', index=False)
    except Exception:
        pass  # position merge failure is non-fatal; filter falls back gracefully


def _merge_ear_positions(raw_folder, ear_summary_path):
    '''Read ears_all_centers.xlsx from the raw LabGym folder and write
    Ear0X / Ear0Y / Ear1X / Ear1Y columns into ear_summary.xlsx.

    ears_all_centers.xlsx has rows like:
        time/ID  |  0          |  1
        0.03     |  (736, 748) |  (812, 750)

    Column 0 → Ear0X/Ear0Y, Column 1 → Ear1X/Ear1Y.
    '''
    centers_path = Path(raw_folder) / 'ears_all_centers.xlsx'
    if not centers_path.exists() or not ear_summary_path.exists():
        return
    try:
        centers = pd.read_excel(centers_path)

        def _parse(val):
            if pd.isna(val):
                return np.nan, np.nan
            s = str(val).strip().lstrip('(').rstrip(')')
            parts = s.split(',')
            if len(parts) != 2:
                return np.nan, np.nan
            try:
                return float(parts[0].strip()), float(parts[1].strip())
            except ValueError:
                return np.nan, np.nan

        parsed0 = [_parse(v) for v in centers.iloc[:, 1]]
        parsed1 = [_parse(v) for v in centers.iloc[:, 2]] if centers.shape[1] > 2 else [(np.nan, np.nan)] * len(parsed0)
        x0 = [p[0] for p in parsed0]; y0 = [p[1] for p in parsed0]
        x1 = [p[0] for p in parsed1]; y1 = [p[1] for p in parsed1]

        summary = pd.read_excel(ear_summary_path, sheet_name='Summary')
        n_sum = len(summary)
        for lst in (x0, y0, x1, y1):
            while len(lst) < n_sum:
                lst.append(np.nan)
        summary['Ear0X'] = x0[:n_sum]; summary['Ear0Y'] = y0[:n_sum]
        summary['Ear1X'] = x1[:n_sum]; summary['Ear1Y'] = y1[:n_sum]

        with pd.ExcelWriter(str(ear_summary_path), engine='openpyxl') as w:
            summary.to_excel(w, sheet_name='Summary', index=False)
    except Exception:
        pass


def _merge_eye_positions(raw_folder, eye_summary_path):
    '''Read eyes_all_centers.xlsx from the raw LabGym folder and write
    Eye0X / Eye0Y / Eye1X / Eye1Y columns into eye_summary.xlsx.

    eyes_all_centers.xlsx has rows like:
        time/ID  |  0          |  1
        0.03     |  (736, 748) |  (812, 750)

    Column 0 → Eye0X/Eye0Y, Column 1 → Eye1X/Eye1Y.
    '''
    centers_path = Path(raw_folder) / 'eyes_all_centers.xlsx'
    if not centers_path.exists() or not eye_summary_path.exists():
        return
    try:
        centers = pd.read_excel(centers_path)

        def _parse(val):
            if pd.isna(val):
                return np.nan, np.nan
            s = str(val).strip().lstrip('(').rstrip(')')
            parts = s.split(',')
            if len(parts) != 2:
                return np.nan, np.nan
            try:
                return float(parts[0].strip()), float(parts[1].strip())
            except ValueError:
                return np.nan, np.nan

        parsed0 = [_parse(v) for v in centers.iloc[:, 1]]
        parsed1 = [_parse(v) for v in centers.iloc[:, 2]] if centers.shape[1] > 2 else [(np.nan, np.nan)] * len(parsed0)
        x0 = [p[0] for p in parsed0]; y0 = [p[1] for p in parsed0]
        x1 = [p[0] for p in parsed1]; y1 = [p[1] for p in parsed1]

        summary = pd.read_excel(eye_summary_path, sheet_name='Summary')
        n_sum = len(summary)
        for lst in (x0, y0, x1, y1):
            while len(lst) < n_sum:
                lst.append(np.nan)
        summary['Eye0X'] = x0[:n_sum]; summary['Eye0Y'] = y0[:n_sum]
        summary['Eye1X'] = x1[:n_sum]; summary['Eye1Y'] = y1[:n_sum]

        with pd.ExcelWriter(str(eye_summary_path), engine='openpyxl') as w:
            summary.to_excel(w, sheet_name='Summary', index=False)
    except Exception:
        pass


def _summary_to_intensity_array(df, event_mask, intensity_cols):
    '''
    Per-frame intensity: mean of intensity_cols for event-masked rows, NaN elsewhere.
    '''
    result = np.full(len(df), np.nan)
    avail = [c for c in intensity_cols if c in df.columns]
    if not avail or not event_mask.any():
        return result
    with warnings.catch_warnings():
        warnings.simplefilter('ignore', RuntimeWarning)
        vals = df.loc[event_mask, avail].apply(pd.to_numeric, errors='coerce').values
        if vals.ndim == 1:
            result[event_mask.values] = vals
        else:
            result[event_mask.values] = np.nanmean(vals, axis=1)
    return result


_CACHE_FILE = '_labgrymace_cache.npz'
_CACHE_VERSION = 20  # bump when filter logic changes to force cache rebuild
                     # (19: added apply_mirror_filter flag to the cache)


def _cache_is_valid(folder, apply_mirror_filter=True):
    '''Return True if the .npz cache exists, matches the current version,
    was built with the same apply_mirror_filter setting, and is newer than all
    three source xlsx files.'''
    cache_path = folder / _CACHE_FILE
    if not cache_path.exists():
        return False
    try:
        with np.load(cache_path, allow_pickle=True) as c:
            if int(c.get('version', 0)) != _CACHE_VERSION:
                return False
            # cache built with a different mirror-filter setting is not reusable
            if bool(c.get('apply_mirror_filter', True)) != bool(apply_mirror_filter):
                return False
        cache_mtime = cache_path.stat().st_mtime
        for xlsx in ('ear_summary.xlsx', 'eye_summary.xlsx', 'nose_summary.xlsx'):
            src = folder / xlsx
            if src.exists() and src.stat().st_mtime > cache_mtime:
                return False
        return True
    except Exception:
        return False


def _load_from_cache(folder):
    '''Load pre-filtered arrays from the .npz cache.'''
    with np.load(folder / _CACHE_FILE, allow_pickle=True) as c:
        return {
            'time':                c['time'],
            'ear_intensity':       c['ear_intensity'],
            'eye_intensity':       c['eye_intensity'],
            'nose_intensity':      c['nose_intensity'],
            'ear0_mirror_removed': c['ear0_mirror_removed'],
            'ear1_mirror_removed': c['ear1_mirror_removed'],
            'eye0_mirror_removed': c['eye0_mirror_removed'],
            'eye1_mirror_removed': c['eye1_mirror_removed'],
            'ear_mirror_removed':  c['ear_mirror_removed'],
            'eye_mirror_removed':  c['eye_mirror_removed'],
            'nose_mirror_removed': c['nose_mirror_removed'],
            'n_frames':            int(c['n_frames']),
            'filter_source':       str(c['filter_source']) + ' [cached]',
        }


def _save_to_cache(folder, result, apply_mirror_filter=True):
    '''Save filtered arrays to .npz cache for fast future loads.
    Does NOT save if any intensity channel is entirely NaN — that indicates
    a transient read failure and should be retried next time.'''
    ear = result['ear_intensity']
    eye = result['eye_intensity']
    nose = result['nose_intensity']
    if np.all(np.isnan(ear)) or np.all(np.isnan(eye)) or np.all(np.isnan(nose)):
        return   # refuse to cache bad data
    try:
        np.savez_compressed(
            folder / _CACHE_FILE,
            version=np.array(_CACHE_VERSION),
            apply_mirror_filter=np.array(bool(apply_mirror_filter)),
            time=result['time'],
            ear_intensity=ear,
            eye_intensity=eye,
            nose_intensity=nose,
            ear0_mirror_removed=result['ear0_mirror_removed'],
            ear1_mirror_removed=result['ear1_mirror_removed'],
            eye0_mirror_removed=result['eye0_mirror_removed'],
            eye1_mirror_removed=result['eye1_mirror_removed'],
            ear_mirror_removed=result['ear_mirror_removed'],
            eye_mirror_removed=result['eye_mirror_removed'],
            nose_mirror_removed=result['nose_mirror_removed'],
            n_frames=np.array(result['n_frames']),
            filter_source=np.array(result['filter_source']),
        )
    except Exception:
        pass   # cache write failure is non-fatal


def _find_mirror_point(hv_x, hv_y, radius):
    '''Find the center of the densest spatial cluster among high-velocity nose
    positions.  Uses a simple O(n²) neighbourhood scan — the number of
    high-velocity frames is at most a few thousand, so this is fast enough.

    Returns (cx, cy, cluster_fraction) where cluster_fraction is the proportion
    of high-vel frames within `radius` of the found centre.
    '''
    if len(hv_x) == 0:
        return 0.0, 0.0, 0.0
    best_count = 0
    best_cx = float(np.mean(hv_x))
    best_cy = float(np.mean(hv_y))
    for xi, yi in zip(hv_x, hv_y):
        dists = np.sqrt((hv_x - xi) ** 2 + (hv_y - yi) ** 2)
        in_r  = dists < radius
        cnt   = int(in_r.sum())
        if cnt > best_count:
            best_count = cnt
            best_cx = float(np.mean(hv_x[in_r]))
            best_cy = float(np.mean(hv_y[in_r]))
    return best_cx, best_cy, best_count / len(hv_x)


def _temporal_cluster_mask(n_frames, hv_mask, window, min_neighbors):
    '''Return a boolean array of length n_frames.  True at position i means
    frame i is a high-velocity frame that has at least min_neighbors other
    high-velocity frames within the ±window neighbourhood — i.e., it is part
    of a recurring burst rather than an isolated fast movement.
    '''
    result  = np.zeros(n_frames, dtype=bool)
    hv_idx  = np.where(hv_mask)[0]
    if len(hv_idx) == 0:
        return result
    for pos in hv_idx:
        neighbours = int(np.sum((hv_idx > pos - window) &
                                (hv_idx < pos + window) &
                                (hv_idx != pos)))
        if neighbours >= min_neighbors:
            result[pos] = True
    return result


def _apply_mirror_filter(arr, vel_arr, mag_arr, x_arr, y_arr, event_mask, channel_label):
    '''Apply the 4-condition mirror-reflection filter to a single tracker's
    intensity array.  Returns (filtered_arr, mode_str, bad_mask).

    Parameters
    ----------
    arr          : intensity array (float, may have NaN)
    vel_arr      : velocity array (float, may have NaN)
    mag_arr      : magnitude_area array (float, may be None or all-NaN)
    x_arr        : X centroid position array (float, may have NaN)
    y_arr        : Y centroid position array (float, may have NaN)
    event_mask   : boolean array — which frames belong to the event type used
                   for the bimodal gate (e.g. NoseBul, EarBL, EyeOT)
    channel_label: short string used in log messages (e.g. "nose", "ear0")

    A frame is removed only when ALL of:
      cond1  velocity > VEL_THRESH
      cond2  position within SPATIAL_RADIUS of the auto-detected mirror point
      cond3  magnitude_area > MAG_THRESH
      cond4  ≥ TEMPORAL_MIN_N other high-vel frames within ±TEMPORAL_WINDOW
    The bimodal pre-check skips the filter entirely for normal recordings.
    '''
    n = len(arr)

    # ── Bimodal gate ──────────────────────────────────────────────────────────
    _ev_vel  = vel_arr[event_mask & ~np.isnan(vel_arr)] if event_mask.any() else np.array([])
    _hv_pct  = float((_ev_vel > VEL_THRESH).mean()) if len(_ev_vel) > 0 else 0.0

    if _hv_pct <= VEL_BIMODAL_PCT:
        return arr.copy(), (f'skipped: {100*_hv_pct:.1f}% {channel_label}>vel '
                            f'(unimodal — no artifact detected)'), np.zeros(n, dtype=bool)

    # ── Condition 1: velocity ─────────────────────────────────────────────────
    cond1 = (vel_arr > VEL_THRESH) & ~np.isnan(vel_arr)

    # ── Condition 2: near fixed mirror point (spatial) ────────────────────────
    _has_pos = ~np.isnan(x_arr) & ~np.isnan(y_arr)
    if _has_pos.any():
        _hv_with_pos = cond1 & _has_pos
        _hv_x = x_arr[_hv_with_pos]
        _hv_y = y_arr[_hv_with_pos]
        if len(_hv_x) > 0:
            _mx, _my, _cfrac = _find_mirror_point(_hv_x, _hv_y, SPATIAL_RADIUS)
            if _cfrac >= SPATIAL_MIN_PCT:
                _dist  = np.sqrt((x_arr - _mx)**2 + (y_arr - _my)**2)
                cond2  = (_dist < SPATIAL_RADIUS) & _has_pos
                sp_note = f'mirror@({_mx:.0f},{_my:.0f}) cluster={100*_cfrac:.0f}%'
            else:
                # Cannot confirm mirror point — do not remove frames
                cond2  = np.zeros(n, dtype=bool)
                sp_note = (f'spatial blocked '
                           f'(cluster={100*_cfrac:.0f}%<{100*SPATIAL_MIN_PCT:.0f}%)')
        else:
            # No high-velocity frames with position — no mirror artifact detectable
            cond2  = np.zeros(n, dtype=bool)
            sp_note = 'spatial blocked (no hv+pos frames)'
    else:
        # No XY position data — spatial condition passes through (use other 3 conditions)
        cond2  = np.ones(n, dtype=bool)
        sp_note = f'spatial skipped (no {channel_label} XY — other 3 conditions apply)'

    # ── Condition 3: large magnitude (displacement / landmark size) ───────────
    if mag_arr is not None and not np.all(np.isnan(mag_arr)):
        cond3    = (mag_arr > MAG_THRESH) & ~np.isnan(mag_arr)
        mag_note = f'mag>{MAG_THRESH}'
    else:
        cond3    = np.ones(n, dtype=bool)
        mag_note = 'mag skipped (no column)'

    # ── Condition 4: temporal recurrence ─────────────────────────────────────
    cond4 = _temporal_cluster_mask(n, cond1, TEMPORAL_WINDOW, TEMPORAL_MIN_N)

    # ── Remove only if ALL 4 satisfied ────────────────────────────────────────
    bad = cond1 & cond2 & cond3 & cond4
    filtered = arr.copy()
    filtered[bad] = np.nan
    mode = (f'engaged: {100*_hv_pct:.1f}%>{VEL_THRESH} (bimodal); '
            f'removed {bad.sum()} frames '
            f'[c1:vel>{VEL_THRESH} | c2:{sp_note} | c3:{mag_note} | '
            f'c4:±{TEMPORAL_WINDOW}fr≥{TEMPORAL_MIN_N}nbr]')
    return filtered, mode, bad


def _load_intensity_from_summary(animal_folder, apply_mirror_filter=True):
    '''
    Read per-frame intensity from ear/eye/nose_summary.xlsx files.

    apply_mirror_filter: when False, the mirror-reflection filter is skipped
    entirely — no frames are removed and the *_mirror_removed masks are all-False.

    On first call the xlsx files are parsed, filtered, and the result is
    cached as _labgrymace_cache.npz in the animal folder.  Subsequent calls
    return the cached arrays directly (10-50x faster than re-reading xlsx).
    The cache is invalidated automatically if any source xlsx is newer.

    Filters applied:
      - Ear:  Ear0Event == BL  (bilateral OR)
      - Eye:  Eye0Event == OT  (bilateral OR)
      - Nose: NoseEvent == Bul AND Velocity <= VEL_THRESH
    Fallback to all-events if any channel < FALLBACK_THRESHOLD_PCT%.

    Returns dict: {time, ear_intensity, eye_intensity, nose_intensity,
                   n_frames, filter_source}.
    '''
    folder = Path(animal_folder)

    # ── Fast path: return cached result ──────────────────────────────────────
    if _cache_is_valid(folder, apply_mirror_filter):
        return _load_from_cache(folder)

    # ── Slow path: parse xlsx files ──────────────────────────────────────────
    def _intensity_array(df, cols):
        avail = [c for c in cols if c in df.columns]
        if not avail:
            return np.full(len(df), np.nan)
        with warnings.catch_warnings():
            warnings.simplefilter('ignore', RuntimeWarning)
            return np.nanmean(df[avail].apply(pd.to_numeric, errors='coerce').values, axis=1)

    def _apply_mask(arr, mask):
        out = arr.copy().astype(float)
        out[~mask] = np.nan
        return out

    # --- Ear ---
    ear_df = pd.read_excel(folder / 'ear_summary.xlsx', sheet_name='Summary')
    time = pd.to_numeric(ear_df.iloc[:, 0], errors='coerce').values
    n = len(time)

    def _col(df, col):
        '''Extract a numeric column; return all-NaN array of length n if absent.'''
        if df is None or col not in df.columns:
            return np.full(n, np.nan)
        return pd.to_numeric(df[col], errors='coerce').values

    ear_int_0 = _intensity_array(ear_df, ['Intensity Area 0'])
    ear_int_1 = _intensity_array(ear_df, ['Intensity Area 1'])
    with warnings.catch_warnings():
        warnings.simplefilter('ignore', RuntimeWarning)
        ear_arr_all = np.nanmean(np.stack([ear_int_0, ear_int_1], axis=1), axis=1)

    # --- Eye ---
    eye_df = None
    eye_int_0 = np.full(n, np.nan)
    eye_int_1 = np.full(n, np.nan)
    try:
        eye_df = pd.read_excel(folder / 'eye_summary.xlsx', sheet_name='Summary')
        if len(eye_df) != n:
            eye_df = None
        else:
            eye_int_0 = _intensity_array(eye_df, ['Intensity Area 0'])
            eye_int_1 = _intensity_array(eye_df, ['Intensity Area 1'])
    except Exception:
        pass
    with warnings.catch_warnings():
        warnings.simplefilter('ignore', RuntimeWarning)
        eye_arr_all = np.nanmean(np.stack([eye_int_0, eye_int_1], axis=1), axis=1)

    # --- Nose ---
    nose_df = None
    try:
        nose_df = pd.read_excel(folder / 'nose_summary.xlsx', sheet_name='Summary')
        nose_arr_all = _intensity_array(nose_df, ['Intensity Area'])
        if len(nose_arr_all) != n:
            nose_df = None
            nose_arr_all = np.full(n, np.nan)
    except Exception:
        nose_arr_all = np.full(n, np.nan)

    # ── 4-condition mirror-reflection filter — Nose ───────────────────────────
    # Per-tracker mirror-removal masks (default all-False; each reassigned if its filter runs)
    bad_nose = bad_ear0 = bad_ear1 = bad_eye0 = bad_eye1 = np.zeros(n, dtype=bool)

    vel_nose_mode = ('mirror filter OFF (opt-in)' if not apply_mirror_filter
                     else 'skipped (no Velocity column)')
    if apply_mirror_filter and nose_df is not None and 'Velocity' in nose_df.columns:
        _bul_mask = ((nose_df['NoseEvent'] == 'Bul').values
                     if 'NoseEvent' in nose_df.columns
                     else np.zeros(n, dtype=bool))
        nose_arr_all, vel_nose_mode, bad_nose = _apply_mirror_filter(
            nose_arr_all,
            _col(nose_df, 'Velocity'),
            _col(nose_df, 'Magnitude Area'),
            _col(nose_df, 'NoseX'),
            _col(nose_df, 'NoseY'),
            _bul_mask,
            'nose')

    # ── 4-condition mirror-reflection filter — Ear tracker 0 & 1 ─────────────
    vel_ear0_mode = ('mirror filter OFF (opt-in)' if not apply_mirror_filter
                     else 'skipped (no Velocity 0 column)')
    vel_ear1_mode = ('mirror filter OFF (opt-in)' if not apply_mirror_filter
                     else 'skipped (no Velocity 1 column)')
    if apply_mirror_filter and ear_df is not None and 'Velocity 0' in ear_df.columns:
        _ear0_bl = ((ear_df['Ear0Event'] == 'BL').values
                    if 'Ear0Event' in ear_df.columns
                    else np.zeros(n, dtype=bool))
        ear_int_0, vel_ear0_mode, bad_ear0 = _apply_mirror_filter(
            ear_int_0,
            _col(ear_df, 'Velocity 0'),
            _col(ear_df, 'Magnitude Area 0'),
            _col(ear_df, 'Ear0X'),
            _col(ear_df, 'Ear0Y'),
            _ear0_bl,
            'ear0')
    if apply_mirror_filter and ear_df is not None and 'Velocity 1' in ear_df.columns:
        _ear1_bl = ((ear_df['Ear1Event'] == 'BL').values
                    if 'Ear1Event' in ear_df.columns
                    else np.zeros(n, dtype=bool))
        ear_int_1, vel_ear1_mode, bad_ear1 = _apply_mirror_filter(
            ear_int_1,
            _col(ear_df, 'Velocity 1'),
            _col(ear_df, 'Magnitude Area 1'),
            _col(ear_df, 'Ear1X'),
            _col(ear_df, 'Ear1Y'),
            _ear1_bl,
            'ear1')
    with warnings.catch_warnings():
        warnings.simplefilter('ignore', RuntimeWarning)
        ear_arr_all = np.nanmean(np.stack([ear_int_0, ear_int_1], axis=1), axis=1)

    # ── 4-condition mirror-reflection filter — Eye tracker 0 & 1 ─────────────
    vel_eye0_mode = ('mirror filter OFF (opt-in)' if not apply_mirror_filter
                     else 'skipped (no Velocity 0 column)')
    vel_eye1_mode = ('mirror filter OFF (opt-in)' if not apply_mirror_filter
                     else 'skipped (no Velocity 1 column)')
    if apply_mirror_filter and eye_df is not None and 'Velocity 0' in eye_df.columns:
        _eye0_ot = ((eye_df['Eye0Event'] == 'OT').values
                    if 'Eye0Event' in eye_df.columns
                    else np.zeros(n, dtype=bool))
        eye_int_0, vel_eye0_mode, bad_eye0 = _apply_mirror_filter(
            eye_int_0,
            _col(eye_df, 'Velocity 0'),
            _col(eye_df, 'Magnitude Area 0'),
            _col(eye_df, 'Eye0X'),
            _col(eye_df, 'Eye0Y'),
            _eye0_ot,
            'eye0')
    if apply_mirror_filter and eye_df is not None and 'Velocity 1' in eye_df.columns:
        _eye1_ot = ((eye_df['Eye1Event'] == 'OT').values
                    if 'Eye1Event' in eye_df.columns
                    else np.zeros(n, dtype=bool))
        eye_int_1, vel_eye1_mode, bad_eye1 = _apply_mirror_filter(
            eye_int_1,
            _col(eye_df, 'Velocity 1'),
            _col(eye_df, 'Magnitude Area 1'),
            _col(eye_df, 'Eye1X'),
            _col(eye_df, 'Eye1Y'),
            _eye1_ot,
            'eye1')
    with warnings.catch_warnings():
        warnings.simplefilter('ignore', RuntimeWarning)
        eye_arr_all = np.nanmean(np.stack([eye_int_0, eye_int_1], axis=1), axis=1)

    # Region-level mirror-removal masks (a region frame is "removed" if the mirror
    # filter dropped ANY of its trackers at that frame)
    ear_mirror_removed  = bad_ear0 | bad_ear1
    eye_mirror_removed  = bad_eye0 | bad_eye1
    nose_mirror_removed = bad_nose

    vel_filter_mode = (f'nose: {vel_nose_mode} | '
                       f'ear0: {vel_ear0_mode} | ear1: {vel_ear1_mode} | '
                       f'eye0: {vel_eye0_mode} | eye1: {vel_eye1_mode}')

    # --- Build event masks ---
    has_ear_events  = any(c in ear_df.columns for c in ['Ear0Event', 'Ear1Event'])
    has_eye_events  = eye_df  is not None and any(c in eye_df.columns  for c in ['Eye0Event', 'Eye1Event'])
    has_nose_events = nose_df is not None and 'NoseEvent' in nose_df.columns

    if not USE_EVENT_FILTER:
        # All-events mode: skip BL/OT/Bul masking, use mirror-filtered arrays directly
        ear_arr, eye_arr, nose_arr = ear_arr_all, eye_arr_all, nose_arr_all
        filter_source = f'all_events+mirror_filter | vel filter: {vel_filter_mode}'
    elif has_ear_events and has_eye_events and has_nose_events:
        ear_mask = np.zeros(n, dtype=bool)
        if 'Ear0Event' in ear_df.columns:
            ear_mask |= (ear_df['Ear0Event'] == 'BL').values
        if 'Ear1Event' in ear_df.columns:
            ear_mask |= (ear_df['Ear1Event'] == 'BL').values

        eye_mask = np.zeros(n, dtype=bool)
        if 'Eye0Event' in eye_df.columns:
            eye_mask |= (eye_df['Eye0Event'] == 'OT').values
        if 'Eye1Event' in eye_df.columns:
            eye_mask |= (eye_df['Eye1Event'] == 'OT').values

        # NoseBul event mask only — velocity already removed from nose_arr_all above
        nose_mask = (nose_df['NoseEvent'] == 'Bul').values

        ear_pct  = ear_mask.sum()  / n * 100
        eye_pct  = eye_mask.sum()  / n * 100
        nose_pct = nose_mask.sum() / n * 100

        # Eye and Nose always use OT / Bul — no fallback
        eye_arr  = _apply_mask(eye_arr_all,  eye_mask)
        nose_arr = _apply_mask(nose_arr_all, nose_mask)

        # Ear: use EarBL if >= threshold, otherwise fall back to EarPB
        if ear_pct >= FALLBACK_THRESHOLD_PCT:
            ear_arr = _apply_mask(ear_arr_all, ear_mask)
            ear_fb  = 'EarBL'
        else:
            ear_pb_mask = np.zeros(n, dtype=bool)
            if 'Ear0Event' in ear_df.columns:
                ear_pb_mask |= (ear_df['Ear0Event'] == 'PB').values
            if 'Ear1Event' in ear_df.columns:
                ear_pb_mask |= (ear_df['Ear1Event'] == 'PB').values
            ear_arr = _apply_mask(ear_arr_all, ear_pb_mask)
            ear_fb  = f'EarPB (fallback: EarBL={ear_pct:.1f}%<{FALLBACK_THRESHOLD_PCT}%)'

        filter_source = (f'filtered ({ear_fb}, eye={eye_pct:.1f}%OT, '
                         f'nose={nose_pct:.1f}%Bul) | '
                         f'vel filter: {vel_filter_mode}')
    else:
        ear_arr, eye_arr, nose_arr = ear_arr_all, eye_arr_all, nose_arr_all
        filter_source = f'all_events (event columns not found) | vel filter: {vel_filter_mode}'

    result = {
        'time':                time,
        'ear_intensity':       ear_arr,
        'eye_intensity':       eye_arr,
        'nose_intensity':      nose_arr,
        'ear0_mirror_removed': bad_ear0,
        'ear1_mirror_removed': bad_ear1,
        'eye0_mirror_removed': bad_eye0,
        'eye1_mirror_removed': bad_eye1,
        'ear_mirror_removed':  ear_mirror_removed,
        'eye_mirror_removed':  eye_mirror_removed,
        'nose_mirror_removed': nose_mirror_removed,
        'n_frames':            n,
        'filter_source':       filter_source,
    }
    _save_to_cache(folder, result, apply_mirror_filter)
    return result


def load_raw_data(folder_path, apply_mirror_filter=None):
    '''
    Load EarBL / EyeOT / NoseBul intensity from pre-generated summary files.

    Expects ear_summary.xlsx, eye_summary.xlsx, nose_summary.xlsx to already
    exist in folder_path (use "Generate Summary Files" first if they don't).

    apply_mirror_filter: opt-in mirror-reflection filter.  When None (default),
    falls back to the module-level APPLY_MIRROR_FILTER (default OFF).  Set True
    only when the LabGym tracker may have picked up mirror reflections.  The
    Figure 4E / Figure 6 repro scripts pass apply_mirror_filter=True explicitly
    so they keep reproducing the published values.

    Returns dict with keys: time, ear_intensity, eye_intensity, nose_intensity,
    n_frames.  Returns None on failure.
    '''
    if apply_mirror_filter is None:
        apply_mirror_filter = APPLY_MIRROR_FILTER
    try:
        return _load_intensity_from_summary(folder_path, apply_mirror_filter)
    except Exception as e:
        import traceback
        print(f'Error loading {folder_path}: {e}\n{traceback.format_exc()}')
        return None


# ============================================================================
# Pain score algorithm
# ============================================================================

def compute_windowed_pain_scores(data, window_size=FRAME_WINDOW):
    '''
    Divide frame-by-frame intensity data into windows and compute pain score
    per window using the per-frame temporal method (lookback=60, 10%-trimmed mean,
    channel renormalization by w_sum).

    Pain score per window = nanmean of compute_per_frame_pain_scores within
    that window's frames.  This is consistent with compute_overall_pain_score
    and with how the batch analysis scripts calculate individual animal scores.

    Score 0   = true baseline (6-mouse avg: 9F1/9F2/9M1/9M3/Baseline_4111/Baseline_4116)
    Score 100 = 1 mg/kg CNO (4-mouse avg: 4M1_clip1/4M1_clip2/5F1/M4)
    Values can legitimately exceed this range.

    Remainder frames are kept as a final partial window so no data is discarded.

    Returns list of dicts with keys:
        window, n_frames_in_window, partial,
        time_start, time_end, time_mid,
        ear_intensity, eye_intensity, nose_intensity,
        Z_ear, Z_eye, Z_nose, raw_score, pain_score
    '''
    # Pre-compute per-frame pain scores for the entire recording
    frame_scores = compute_per_frame_pain_scores(data, lookback=60)

    n = data['n_frames']
    n_full_windows = n // window_size
    remainder = n % window_size
    results = []

    # Determine slice boundaries: full windows + optional partial tail
    slices = [(i * window_size, (i + 1) * window_size, False) for i in range(n_full_windows)]
    if remainder > 0:
        slices.append((n_full_windows * window_size, n, True))

    t = data['time']
    for win_idx, (s, e, is_partial) in enumerate(slices):
        ear  = float(np.nanmean(data['ear_intensity'][s:e]))
        eye  = float(np.nanmean(data['eye_intensity'][s:e]))
        nose = float(np.nanmean(data['nose_intensity'][s:e]))

        # Intensity-mean Z-scores and raw score kept for metadata / reporting
        Z_ear  = (CNO_EAR_MEAN  - ear)  / CNO_EAR_STD
        Z_eye  = (CNO_EYE_MEAN  - eye)  / CNO_EYE_STD
        Z_nose = (CNO_NOSE_MEAN - nose) / CNO_NOSE_STD
        raw    = W_EAR * Z_ear + W_EYE * Z_eye + W_NOSE * Z_nose

        # Pain score: nanmean of per-frame scores within this window
        score = float(np.nanmean(frame_scores[s:e]))

        results.append({
            'window':             win_idx + 1,
            'n_frames_in_window': e - s,
            'partial':            is_partial,
            'time_start':         float(t[s]),
            'time_end':           float(t[e - 1]),
            'time_mid':           float(np.nanmean(t[s:e])),
            'ear_intensity':      ear,
            'eye_intensity':      eye,
            'nose_intensity':     nose,
            'Z_ear':              Z_ear,
            'Z_eye':              Z_eye,
            'Z_nose':             Z_nose,
            'raw_score':          raw,
            'pain_score':         score,
        })

    return results


def compute_overall_pain_score(data):
    '''
    Compute a single pain score for the entire recording using the per-frame
    temporal method: pain_score = nanmean(compute_per_frame_pain_scores(...)).

    This is consistent with compute_windowed_pain_scores and with how batch
    analysis scripts (reprocess_error_animals.py, run_cno_analysis.py) compute
    individual animal scores.

    Z_ear / Z_eye / Z_nose / raw_score are derived from full-recording intensity
    means and kept for metadata / backwards-compatibility with the Summary sheet.

    Returns a dict with keys:
        ear_intensity, eye_intensity, nose_intensity,
        Z_ear, Z_eye, Z_nose, raw_score, pain_score
    '''
    ear  = float(np.nanmean(data['ear_intensity']))
    eye  = float(np.nanmean(data['eye_intensity']))
    nose = float(np.nanmean(data['nose_intensity']))

    # Intensity-mean Z-scores for metadata
    Z_ear  = (CNO_EAR_MEAN  - ear)  / CNO_EAR_STD
    Z_eye  = (CNO_EYE_MEAN  - eye)  / CNO_EYE_STD
    Z_nose = (CNO_NOSE_MEAN - nose) / CNO_NOSE_STD
    raw    = W_EAR * Z_ear + W_EYE * Z_eye + W_NOSE * Z_nose

    # Overall pain score: nanmean of per-frame temporal scores (correct formula)
    frame_scores = compute_per_frame_pain_scores(data, lookback=60)
    score = float(np.nanmean(frame_scores))

    return {
        'ear_intensity':  ear,
        'eye_intensity':  eye,
        'nose_intensity': nose,
        'Z_ear':          Z_ear,
        'Z_eye':          Z_eye,
        'Z_nose':         Z_nose,
        'raw_score':      raw,
        'pain_score':     score,
    }


# ============================================================================
# Per-frame pain score — causal rolling window with trimmed mean
# ============================================================================

def compute_per_frame_pain_scores_detailed(data, lookback=60):
    '''
    Compute the per-frame pain score AND every intermediate quantity, so that a
    reviewer can trace each frame end to end.

    For every frame f the score uses a causal (backward-looking) window of
    `lookback` frames ending at f — a 2-second / 60-frame window at 30 fps.
    Within that window each channel's intensity is reduced by a 10%-trimmed mean
    (drops the top 10% of values, suppressing tracking-artifact frames whose
    intensity is 3x+ the normal range). Then per-channel Z = (CNO_mean - x)/CNO_std,
    a weight-renormalized raw score over the available channels, and a 0-100
    normalization pinned to baseline (0) and 1 mg/kg CNO (100).

    Frames 0..lookback-2 have pain_score NaN (insufficient history).

    Returns a dict of equal-length (n_frames) arrays:
      frame, time_s, win_start_frame, win_end_frame, win_start_s, win_end_s,
      ear_2s_mean, eye_2s_mean, nose_2s_mean, Z_ear, Z_eye, Z_nose,
      raw_score, pain_score
    '''
    ear  = data['ear_intensity']
    eye  = data['eye_intensity']
    nose = data['nose_intensity']
    t    = np.asarray(data['time'], dtype=float)
    n    = data['n_frames']
    lb   = lookback

    def _trimmed_causal_mean(arr, lb, trim_pct=0.10):
        '''Causal rolling mean with top trim_pct of values removed per window.'''
        out = np.full(len(arr), np.nan)
        for f in range(lb - 1, len(arr)):
            window = arr[f - lb + 1: f + 1]
            valid  = window[~np.isnan(window)]
            if len(valid) == 0:
                continue
            # Drop the top trim_pct highest values (likely tracking errors)
            k = max(1, int(len(valid) * (1 - trim_pct)))
            out[f] = np.mean(np.partition(valid, k - 1)[:k])
        return out

    ear_m  = _trimmed_causal_mean(ear,  lb)
    eye_m  = _trimmed_causal_mean(eye,  lb)
    nose_m = _trimmed_causal_mean(nose, lb)

    # Compute per-channel Z-scores — NaN where that channel was not tracked
    Z_ear  = (CNO_EAR_MEAN  - ear_m)  / CNO_EAR_STD
    Z_eye  = (CNO_EYE_MEAN  - eye_m)  / CNO_EYE_STD
    Z_nose = (CNO_NOSE_MEAN - nose_m) / CNO_NOSE_STD

    # Per-frame weighted average over available channels only (renormalize weights)
    ear_ok  = ~np.isnan(Z_ear)
    eye_ok  = ~np.isnan(Z_eye)
    nose_ok = ~np.isnan(Z_nose)
    w_sum = (W_EAR * ear_ok + W_EYE * eye_ok + W_NOSE * nose_ok).astype(float)
    w_sum[w_sum == 0] = np.nan          # all channels missing → NaN

    raw = (np.where(ear_ok,  W_EAR  * Z_ear,  0.0) +
           np.where(eye_ok,  W_EYE  * Z_eye,  0.0) +
           np.where(nose_ok, W_NOSE * Z_nose, 0.0)) / w_sum

    scores = np.where(
        ~np.isnan(raw),
        100.0 * (raw - BASELINE_RAW_SCORE) / (CNO_1MG_RAW_SCORE - BASELINE_RAW_SCORE),
        np.nan,
    )

    # 2 s (lookback-frame) causal window bounds for every frame
    frame     = np.arange(n)
    win_start = np.maximum(0, frame - lb + 1)
    win_end   = frame
    _tidx     = lambda idx: t[np.clip(idx, 0, n - 1)] if n > 0 else np.array([])
    return {
        'frame':           frame,
        'time_s':          t,
        'win_start_frame': win_start,
        'win_end_frame':   win_end,
        'win_start_s':     _tidx(win_start),
        'win_end_s':       _tidx(win_end),
        'ear_intensity_raw':  ear,
        'eye_intensity_raw':  eye,
        'nose_intensity_raw': nose,
        'ear_2s_mean':     ear_m,
        'eye_2s_mean':     eye_m,
        'nose_2s_mean':    nose_m,
        'Z_ear':           Z_ear,
        'Z_eye':           Z_eye,
        'Z_nose':          Z_nose,
        'raw_score':       raw,
        'pain_score':      scores,
    }


# Column order for the per-frame traceability table
_PER_FRAME_COLUMNS = [
    'frame', 'time_s', 'win_start_frame', 'win_end_frame', 'win_start_s', 'win_end_s',
    'ear_intensity_raw', 'eye_intensity_raw', 'nose_intensity_raw',
    'ear_2s_mean', 'eye_2s_mean', 'nose_2s_mean', 'Z_ear', 'Z_eye', 'Z_nose',
    'raw_score', 'pain_score',
]


def per_frame_pain_dataframe(data, lookback=60):
    '''One row per frame with the full 2 s-window breakdown (see
    compute_per_frame_pain_scores_detailed). Full precision — a raw traceability
    table; averaging pain_score within a 3000-frame window reproduces the
    corresponding row of compute_windowed_pain_scores.'''
    d = compute_per_frame_pain_scores_detailed(data, lookback=lookback)
    return pd.DataFrame({c: d[c] for c in _PER_FRAME_COLUMNS})


def compute_per_frame_pain_scores(data, lookback=60):
    '''Per-frame pain score as a numpy array of shape (n_frames,). Thin wrapper
    over compute_per_frame_pain_scores_detailed (returns only the pain_score
    column) — public behaviour is unchanged.'''
    return compute_per_frame_pain_scores_detailed(data, lookback=lookback)['pain_score']


# ============================================================================
# Pain-score output writer (shared by GUI button and headless batch driver)
# ============================================================================

_WINDOW_FIELDS = ['pain_score', 'ear_intensity', 'eye_intensity', 'nose_intensity',
                  'Z_ear', 'Z_eye', 'Z_nose', 'raw_score']


def _window_range_summary(windows, w_start, w_end):
    '''Per-field mean over the windows kept in [w_start, w_end] (1-based, inclusive).
    Same-source / same-precision (matches Figure 6): each window field is rounded to 2 dp
    FIRST (the single displayed source), then averaged, then the mean is rounded to 2 dp.
    So averaging the shown 2 dp window rows reproduces this summary exactly, and pain_score
    equals the Figure 6 per-recording value (0-500 s = mean of the 2 dp windows 1..5).'''
    kept = [w for w in windows if w_start <= w['window'] <= w_end]
    out = {'n_windows_used': len(kept)}
    for f in _WINDOW_FIELDS:
        vals = [round(w[f], 2) for w in kept if np.isfinite(w.get(f, np.nan))]
        out[f] = round(float(np.mean(vals)), 2) if vals else float('nan')
    return out


def write_pain_score_outputs(records, output_path, window_range=None):
    '''Write the full pain-score output set for a batch of records to output_path.
    Shared by the GUI "Calculate" button and the headless batch driver so both
    produce byte-identical artifacts:
      - pain_scores.xlsx             (one sheet per animal: per-window rows + Summary)
      - pain_scores_per_frame.xlsx   (one sheet per animal: per-frame 2 s-window breakdown)
      - pain_score_chart.png         (one dot plot; one point per animal)
    `records` = list of {'name', 'folder', 'data', 'windows'}. Returns output paths.

    window_range : (w_start, w_end) 1-based inclusive window indices, or None
        None (default, GUI full recording) -> every window/frame is written; the Summary
        overall + the mean bar use the WHOLE recording.
        (w_start, w_end) -> the output is RESTRICTED to those windows (Figure 6 uses (1, 5)
        = 0-500 s). Then EVERYTHING is consistent over that range: the per-window sheets
        contain only the kept windows, the per-frame sheets only the kept frames, the
        Summary reports the per-field mean of the kept windows (so averaging the shown
        window rows reproduces it, and pain_score matches Figure 6), and the plot is drawn
        over the kept range. A YELLOW note cell states the range in every sheet.
    '''
    from openpyxl.styles import PatternFill, Font
    YELLOW = PatternFill('solid', fgColor='FFFF00')
    BOLD   = Font(bold=True)

    wr = window_range  # (w_start, w_end) inclusive, 1-based; or None

    def kept(windows):
        return list(windows) if wr is None else [w for w in windows
                                                 if wr[0] <= w['window'] <= wr[1]]

    # human range label + note (only in restricted mode).
    # Derive from the REQUESTED bounds (deterministic 100 s @30fps windows), NOT from the
    # first record, so a heterogeneous batch cannot mislabel later animals.
    rng_lbl = ''
    note = None
    if wr is not None:
        sec_per_win = FRAME_WINDOW / 30.0
        s0, e1 = (wr[0] - 1) * sec_per_win, wr[1] * sec_per_win
        rng_lbl = f'{s0:.0f}-{e1:.0f} s'
        note = (f'RESTRICTED to {rng_lbl} (windows {wr[0]}-{wr[1]}) = the analysis window '
                f'used in Figure 6. Every row and number in this file is computed '
                f'over {rng_lbl} ONLY; later time is excluded.')

    # ── Excel (per-window) ─────────────────────────────────────────────────
    xlsx_path = os.path.join(output_path, 'pain_scores.xlsx')
    writer = pd.ExcelWriter(xlsx_path, engine='openpyxl')
    summary_rows = []
    for rec in records:
        kw = kept(rec['windows'])
        df = pd.DataFrame(kw)
        if wr is not None:                       # 2 dp single source (matches Figure 6)
            for f in _WINDOW_FIELDS:
                if f in df.columns:
                    df[f] = df[f].round(2)
        df.to_excel(writer, sheet_name=rec['name'][:31], index=False)
        ws = writer.sheets[rec['name'][:31]]
        if wr is None:
            overall = compute_overall_pain_score(rec['data'])
            srow = {
                'animal':                 rec['name'],
                'n_windows':              len(kw),
                'overall_pain_score':     round(overall['pain_score'], 4),
                'overall_ear_intensity':  round(overall['ear_intensity'], 4),
                'overall_eye_intensity':  round(overall['eye_intensity'], 4),
                'overall_nose_intensity': round(overall['nose_intensity'], 4),
                'overall_Z_ear':          round(overall['Z_ear'], 4),
                'overall_Z_eye':          round(overall['Z_eye'], 4),
                'overall_Z_nose':         round(overall['Z_nose'], 4),
                'overall_raw_score':      round(overall['raw_score'], 4),
            }
            avg_label = 'Overall (whole recording)'
            avg_map = {'pain_score': 'overall_pain_score', 'ear_intensity': 'overall_ear_intensity',
                       'eye_intensity': 'overall_eye_intensity', 'nose_intensity': 'overall_nose_intensity',
                       'Z_ear': 'overall_Z_ear', 'Z_eye': 'overall_Z_eye', 'Z_nose': 'overall_Z_nose',
                       'raw_score': 'overall_raw_score'}
        else:
            s = _window_range_summary(rec['windows'], wr[0], wr[1])
            srow = {
                'animal':          rec['name'],
                'time_range':      rng_lbl,
                'n_windows_used':  s['n_windows_used'],
                'pain_score':      round(s['pain_score'], 4),
                'ear_intensity':   round(s['ear_intensity'], 4),
                'eye_intensity':   round(s['eye_intensity'], 4),
                'nose_intensity':  round(s['nose_intensity'], 4),
                'Z_ear':           round(s['Z_ear'], 4),
                'Z_eye':           round(s['Z_eye'], 4),
                'Z_nose':          round(s['Z_nose'], 4),
                'raw_score':       round(s['raw_score'], 4),
            }
            avg_label = f'Average ({rng_lbl})'
            avg_map = {k: k for k in ('pain_score', 'ear_intensity', 'eye_intensity',
                       'nose_intensity', 'Z_ear', 'Z_eye', 'Z_nose', 'raw_score')}
        summary_rows.append(srow)

        # Bottom row on the animal's own tab = its Summary row (SAME source dict, so
        # the numbers are identical), so the mean pain score is visible without opening
        # Summary. The label goes in the 'window' column (a string, not a window index),
        # so downstream window-row parsers skip it. In restricted mode the later block
        # applies the 0.00 format + yellow pain_score fill to this row too.
        hdr = [c.value for c in ws[1]]
        arow = ws.max_row + 1
        lc = (hdr.index('window') + 1) if 'window' in hdr else 1
        ws.cell(row=arow, column=lc, value=avg_label).font = BOLD
        for col, key in avg_map.items():
            if col in hdr:
                ws.cell(row=arow, column=hdr.index(col) + 1, value=srow[key]).font = BOLD
    pd.DataFrame(summary_rows).to_excel(writer, sheet_name='Summary', index=False)

    # restricted mode: uniform 2-dp display ('0.00'), yellow note, highlight pain_score
    if wr is not None:
        FMT = '0.00'
        win_fmt = set(_WINDOW_FIELDS) | {'time_start', 'time_end', 'time_mid'}

        def _fmt_cols(ws, names):
            '''Force '0.00' on every column whose header is in *names*; return the header list.'''
            hdr = [ws.cell(row=1, column=c).value for c in range(1, ws.max_column + 1)]
            for c, name in enumerate(hdr, start=1):
                if name in names:
                    for r in range(2, ws.max_row + 1):
                        ws.cell(row=r, column=c).number_format = FMT
            return hdr

        for rec in records:
            ws = writer.sheets[rec['name'][:31]]
            hdr = _fmt_cols(ws, win_fmt)
            if 'pain_score' in hdr:
                pcol = hdr.index('pain_score') + 1
                for r in range(1, ws.max_row + 1):
                    ws.cell(row=r, column=pcol).fill = YELLOW
            nc = ws.cell(row=1, column=ws.max_column + 2, value=note)
            nc.font = BOLD; nc.fill = YELLOW
        sws = writer.sheets['Summary']
        hdr = _fmt_cols(sws, set(_WINDOW_FIELDS))
        if 'pain_score' in hdr:
            pcol = hdr.index('pain_score') + 1
            for r in range(1, sws.max_row + 1):
                sws.cell(row=r, column=pcol).fill = YELLOW
            sws.cell(row=1, column=pcol).font = BOLD
        snc = sws.cell(row=1, column=sws.max_column + 2,
                       value=note + ' `pain_score` (yellow) = mean of the shown window '
                             'rows = the value plotted in Figure 6.')
        snc.font = BOLD; snc.fill = YELLOW

    # Put Summary first so it's the tab users land on (per-animal averages up
    # front, no scrolling to the last tab). GUI + driver share this writer.
    wb = writer.book
    wb.move_sheet('Summary', offset=-wb.sheetnames.index('Summary'))
    wb.active = 0
    writer.close()

    # ── Per-frame pain score (2 s / 60-frame causal window) ────────────────
    ORANGE = PatternFill('solid', fgColor='FFC000')
    ORANGE_LEGEND = (
        'ORANGE = the mirror-reflection filter removed THIS ORGAN (ear / eye / nose) at this '
        'frame. Removal is PER ORGAN, NOT per frame: only the affected organ value is dropped; '
        'the other organs still count and this frame pain score is recomputed from the remaining '
        'organs (a frame is fully excluded only if all three organs are removed at once). '
        'blank + orange = the organ value is fully gone here (nose removed, or BOTH ear/eye '
        'trackers). value + orange = only one of the organ two trackers (the reflection) was '
        'dropped, so the organ value is recomputed from the surviving tracker.')
    PF_NUMFMT = '0.00'   # uniform 2-dp DISPLAY; stored values stay full precision (== actual)
    PF_INT_COLS = ('frame', 'win_start_frame', 'win_end_frame')
    _RAW_MASK = {'ear_intensity_raw': 'ear_mirror_removed',
                 'eye_intensity_raw': 'eye_mirror_removed',
                 'nose_intensity_raw': 'nose_mirror_removed'}
    pf_path = os.path.join(output_path, 'pain_scores_per_frame.xlsx')
    pf_writer = pd.ExcelWriter(pf_path, engine='openpyxl')
    for rec in records:
        pdf = per_frame_pain_dataframe(rec['data'])
        if wr is not None:                                    # keep only frames in range
            f0, f1 = (wr[0] - 1) * FRAME_WINDOW, wr[1] * FRAME_WINDOW
            pdf = pdf[(pdf['frame'] >= f0) & (pdf['frame'] < f1)]
        pdf.to_excel(pf_writer, sheet_name=rec['name'][:31], index=False)
        ws = pf_writer.sheets[rec['name'][:31]]
        hdr = [ws.cell(row=1, column=c).value for c in range(1, ws.max_column + 1)]
        n_data_cols = len(hdr)
        frames = pdf['frame'].to_numpy()
        note_col = n_data_cols + 2
        # orange-fill the frames the mirror filter removed, per region
        for raw_col, mask_key in _RAW_MASK.items():
            if raw_col not in hdr or mask_key not in rec['data']:
                continue
            col = hdr.index(raw_col) + 1
            mask = np.asarray(rec['data'][mask_key], dtype=bool)
            for i, f in enumerate(frames):
                if f < len(mask) and mask[f]:
                    ws.cell(row=i + 2, column=col).fill = ORANGE
        # uniform 2-dp display on every numeric (non-integer) data column
        int_idx = {hdr.index(c) for c in PF_INT_COLS if c in hdr}
        for row in ws.iter_rows(min_row=2, max_col=n_data_cols):
            for j, cell in enumerate(row):
                if j not in int_idx and cell.value is not None:
                    cell.number_format = PF_NUMFMT
        lc = ws.cell(row=1, column=note_col, value=ORANGE_LEGEND)
        lc.font = BOLD; lc.fill = ORANGE
        if wr is not None:                                    # yellow range note (row 2), pain_score header
            rc = ws.cell(row=2, column=note_col,
                         value=note + ' Per-frame values are shown to 2 dp; the stored value '
                               'is full precision (equals the actual computed value).')
            rc.font = BOLD; rc.fill = YELLOW
            if 'pain_score' in hdr:
                ws.cell(row=1, column=hdr.index('pain_score') + 1).fill = YELLOW
    pf_writer.close()

    # ── One dot plot: every animal is a single point ───────────────────────
    # One figure for the whole batch, one point per animal -- deliberately a
    # dot plot rather than bars, so individual animals stay visible.
    names  = [rec['name'] for rec in records]
    if wr is None:
        scores = [compute_overall_pain_score(rec['data'])['pain_score'] for rec in records]
        ylabel = 'Overall Pain Score (full-recording mean)'
    else:
        scores = [_window_range_summary(rec['windows'], wr[0], wr[1])['pain_score']
                  for rec in records]
        ylabel = f'Pain Score ({rng_lbl} mean)'

    n_animals = len(records)
    fig, ax = plt.subplots(figsize=(max(3.2, 0.55 * n_animals + 2.0), 3.8))
    x = np.arange(1, n_animals + 1)
    ax.axhline(0, color='gray', linestyle='--', linewidth=1.2, alpha=0.6, zorder=1,
               label='Baseline (0)')
    ax.axhline(100, color='red', linestyle='--', linewidth=1.2, alpha=0.4, zorder=1,
               label='1 mg/kg CNO (100)')
    ax.plot(x, scores, linestyle='none', marker='o', markersize=9, zorder=3,
            color='#1f77b4', markeredgecolor='white', markeredgewidth=0.8)
    for xi, val in zip(x, scores):
        if np.isfinite(val):
            ax.annotate(f'{val:.1f}', (xi, val), textcoords='offset points',
                        xytext=(0, 9), ha='center', fontsize=8)
    for sp in ('top', 'right'):
        ax.spines[sp].set_visible(False)
    for sp in ('left', 'bottom'):
        ax.spines[sp].set_linewidth(2.2)
    ax.set_xticks(x)
    ax.set_xticklabels(names, rotation=30, ha='right', fontsize=9)
    ax.set_ylabel(ylabel, labelpad=3, fontsize=11)
    ax.tick_params(axis='both', length=5, labelsize=10)
    finite = [v for v in scores if np.isfinite(v)]
    lo = min(min(finite), 0) - 10 if finite else -10
    hi = max(max(finite), 100) + 14 if finite else 114
    ax.set_ylim(lo, hi)
    ax.set_xlim(0.4, n_animals + 0.6)
    # Above the axes: the reference lines span the full width so an in-axes
    # legend collides with one of them, and below the axes it collides with the
    # rotated animal labels.
    ax.legend(loc='lower center', bbox_to_anchor=(0.5, 1.01), ncol=2,
              frameon=False, fontsize=9)
    plt.tight_layout()
    chart_path = os.path.join(output_path, 'pain_score_chart.png')
    fig.savefig(chart_path, dpi=300, bbox_inches='tight', facecolor='white')
    plt.close(fig)

    return {'pain_scores': xlsx_path, 'pain_scores_per_frame': pf_path,
            'pain_score_chart': chart_path}


# ============================================================================
# Reviewer traceability: merged-value columns and correlation-input tables
# ============================================================================

# Per summary file: (load_raw_data key, header of the yellow "merged" column).
_MERGED_COL = {
    'ear_summary.xlsx':  ('ear_intensity',  'Ear_merged_intensity(feeds pain score)'),
    'eye_summary.xlsx':  ('eye_intensity',  'Eye_merged_intensity(feeds pain score)'),
    'nose_summary.xlsx': ('nose_intensity', 'Nose_merged_intensity(feeds pain score)'),
}
_MERGED_RULE = (
    'LabGrymace merge rule (per frame, intensity_area channel, mirror filter applied):\n'
    ' - both trackers present -> mean of the two\n'
    ' - only one present        -> that single value (the other was mirror-deleted or missing)\n'
    ' - both missing/deleted     -> blank (NA)\n'
    'This is the per-frame value fed into the pain score (Figure 4E / Figure 6). It spans ALL '
    'frames (not filtered by behavior). YELLOW marks it as the key downstream value.')


def add_pain_score_merged_column(folder, apply_mirror_filter=True):
    '''Write the YELLOW "<organ>_merged_intensity(feeds pain score)" column into the
    ear/eye/nose summary files in `folder`, so reviewers can read the exact per-frame
    value that feeds the pain score.

    The value is the per-frame merged intensity_area (mirror rule: both->mean, one->that
    one, none->blank), i.e. load_raw_data(...)['ear_intensity'] etc. It is NOT behavior
    filtered (spans every frame). apply_mirror_filter defaults to True to match the
    published Figure 4E / Figure 6 convention; pass False for the unfiltered merge.

    Idempotent: reuses the column by header name. Adds no other data. Returns
    {filename: n_values_written} or None if the folder can't be loaded.'''
    import openpyxl
    from openpyxl.styles import PatternFill, Font
    from openpyxl.comments import Comment
    YELLOW = PatternFill('solid', fgColor='FFFF00')
    BOLD = Font(bold=True)
    folder = Path(folder)
    data = load_raw_data(str(folder), apply_mirror_filter=apply_mirror_filter)
    if data is None:
        return None
    out = {}
    for fname, (key, header) in _MERGED_COL.items():
        p = folder / fname
        if not p.exists():
            continue
        merged = np.asarray(data[key], dtype=float)
        wb = openpyxl.load_workbook(str(p))
        ws = wb['Summary'] if 'Summary' in wb.sheetnames else wb[wb.sheetnames[0]]
        col = None
        for c in range(1, ws.max_column + 1):
            if ws.cell(row=1, column=c).value == header:
                col = c
                break
        if col is None:
            col = ws.max_column + 1
        hc = ws.cell(row=1, column=col, value=header)
        hc.font = BOLD
        hc.fill = YELLOW
        hc.comment = Comment(_MERGED_RULE, 'LabGrymace')
        n_val = 0
        last = min(len(merged), ws.max_row - 1)
        for i in range(last):
            cell = ws.cell(row=i + 2, column=col)
            v = merged[i]
            if np.isfinite(v):
                cell.value = float(v)
                cell.number_format = '0.000'
                n_val += 1
            else:
                cell.value = None
            cell.fill = YELLOW
        wb.save(str(p))
        out[fname] = n_val
    return out


# Nine facial parameters used in the Figure-3 correlation (raw, behavior-gated).
_CORR_PARAMS = ['acceleration', 'intensity_area', 'intensity_length', 'magnitude_area',
                'magnitude_length', 'speed', 'velocity', 'vigor_area', 'vigor_length']

# organ -> (summary file, bilateral, event columns, painful event, resting event,
#           load_raw_data intensity key, per-tracker mirror-mask keys, column prefix)
_PANEL_ORGANS = {
    'ear':  ('ear_summary.xlsx',  True,  ('Ear0Event', 'Ear1Event'), 'PB',  'BL',
             'ear_intensity',  ('ear0_mirror_removed', 'ear1_mirror_removed'), 'Ear'),
    'eye':  ('eye_summary.xlsx',  True,  ('Eye0Event', 'Eye1Event'), 'OT',  'BL',
             'eye_intensity',  ('eye0_mirror_removed', 'eye1_mirror_removed'), 'Eye'),
    'nose': ('nose_summary.xlsx', False, ('NoseEvent',),             'Bul', 'BL',
             'nose_intensity', ('nose_mirror_removed',),             'Nose'),
}
_GREEN = '92D050'


def add_behavior_and_painscore_panels(folder, apply_mirror_filter=True):
    '''Write the two reviewer subsheets into each summary file in `folder`:

      PanelB_behavior  - per frame: the two sides' behavior categories, a Group label
                         (pain / not-pain / both / none; "both" = the two sides carry
                         different categories, highlighted green) and the merged value of
                         each of the nine parameters (both sides -> mean, one -> that
                         side, none -> blank). This is the input to the correlation and
                         dose-response analyses.
      PanelC_painscore - per frame: each side's mirror-filtered intensity_area (removed
                         frames orange), the merged value that feeds the pain score
                         (yellow header) and per-side removal flags. All frames are used;
                         no behavior gating is applied here.

    Sheets are placed directly after "Summary" and rebuilt on each call (idempotent).'''
    import openpyxl
    from openpyxl.styles import PatternFill, Font
    BOLD = Font(bold=True)
    GREEN = PatternFill('solid', fgColor=_GREEN)
    ORANGE = PatternFill('solid', fgColor='FFC000')
    YELLOW = PatternFill('solid', fgColor='FFFF00')
    folder = Path(folder)
    data = load_raw_data(str(folder), apply_mirror_filter=apply_mirror_filter)
    if data is None:
        return None

    def _num(x):
        return None if (x is None or not np.isfinite(x)) else round(float(x), 4)

    written = 0
    for organ, (fname, bilat, evc, pain, rest, key, maskkeys, pfx) in _PANEL_ORGANS.items():
        p = folder / fname
        if not p.exists():
            continue
        df = pd.read_excel(p, sheet_name='Summary')
        n = len(df)
        time = df['Time'].values if 'Time' in df.columns else np.arange(n)

        # merged value per parameter (same rule for every parameter)
        merged = {}
        for prm in _CORR_PARAMS:
            if bilat:
                d0 = pd.to_numeric(df[_corr_col(prm) + ' 0'], errors='coerce').values
                d1 = pd.to_numeric(df[_corr_col(prm) + ' 1'], errors='coerce').values
                with warnings.catch_warnings():
                    warnings.simplefilter('ignore', RuntimeWarning)
                    merged[prm] = np.nanmean(np.stack([d0, d1], axis=1), axis=1)
            else:
                merged[prm] = pd.to_numeric(df[_corr_col(prm)], errors='coerce').values

        # behavior grouping from the two sides
        e0 = df[evc[0]].astype(object).values
        e1 = df[evc[1]].astype(object).values if bilat else None
        is_p = (e0 == pain) | (e1 == pain) if bilat else (e0 == pain)
        is_r = (e0 == rest) | (e1 == rest) if bilat else (e0 == rest)
        group = np.where(is_p & is_r, 'both',
                         np.where(is_p, 'pain', np.where(is_r, 'not-pain', 'none')))

        # per-side mirror-filtered intensity_area
        sides = []
        for i, mk in enumerate(maskkeys):
            col = f'Intensity Area {i}' if bilat else 'Intensity Area'
            raw = pd.to_numeric(df[col], errors='coerce').values.astype(float)
            bad = np.asarray(data.get(mk, np.zeros(n, dtype=bool)), dtype=bool)
            filt = raw.copy()
            m = min(len(filt), len(bad))
            filt[:m][bad[:m]] = np.nan
            sides.append((filt, bad))
        dmerged = np.asarray(data[key], dtype=float)

        wb = openpyxl.load_workbook(str(p))
        for sh in ('PanelB_behavior', 'PanelC_painscore'):
            if sh in wb.sheetnames:
                del wb[sh]

        ws = wb.create_sheet('PanelB_behavior')
        hdr = ([ 'Time', f'{pfx}0Event', f'{pfx}1Event', 'Group'] if bilat
               else ['Time', f'{pfx}Event', 'Group']) + _CORR_PARAMS
        ws.append(hdr)
        for c in ws[1]:
            c.font = BOLD
        gcol = hdr.index('Group') + 1
        for i in range(n):
            row = ([time[i], e0[i], e1[i], group[i]] if bilat else [time[i], e0[i], group[i]])
            row += [_num(merged[prm][i]) for prm in _CORR_PARAMS]
            ws.append(row)
        for i in range(n):
            if group[i] == 'both':
                ws.cell(row=i + 2, column=gcol).fill = GREEN

        ws2 = wb.create_sheet('PanelC_painscore')
        if bilat:
            hdr2 = ['Time', f'IntensityArea_{pfx}0_filtered', f'IntensityArea_{pfx}1_filtered',
                    'merged_intensity_area(pain score input)', f'{pfx}0_removed', f'{pfx}1_removed']
        else:
            hdr2 = ['Time', f'IntensityArea_{pfx}_filtered',
                    'merged_intensity_area(pain score input)', f'{pfx}_removed']
        ws2.append(hdr2)
        for c in ws2[1]:
            c.font = BOLD
        mcol = hdr2.index('merged_intensity_area(pain score input)') + 1
        ws2.cell(row=1, column=mcol).fill = YELLOW
        for i in range(n):
            row = [time[i]] + [_num(s[0][i]) for s in sides]
            row += [_num(dmerged[i]) if i < len(dmerged) else None]
            row += ['YES' if (i < len(s[1]) and s[1][i]) else '' for s in sides]
            ws2.append(row)
        for i in range(n):
            for j, (_, bad) in enumerate(sides):
                if i < len(bad) and bad[i]:
                    ws2.cell(row=i + 2, column=2 + j).fill = ORANGE

        order = [wb[s] for s in ('Summary', 'PanelB_behavior', 'PanelC_painscore')
                 if s in wb.sheetnames]
        for sheet in wb.worksheets:
            if sheet not in order:
                order.append(sheet)
        wb._sheets = order
        tmp = str(p) + '.writing.xlsx'
        wb.save(tmp)
        os.replace(tmp, str(p))
        wb.close()
        written += 1
    return written


def _corr_col(p):
    '''Parameter key -> summary column base name ('intensity_area' -> 'Intensity Area').'''
    return p.replace('_', ' ').title()


def _correlation_heatmap(corr, out_png, title=None):
    '''Figure-3 style Pearson-correlation heatmap (RdBu_r, white gridlines, 2-dp cell
    values, "Pearson Correlation" colorbar). corr = a square DataFrame. Writes out_png.'''
    import matplotlib.colors as mcolors
    plt.rcParams.update({'axes.linewidth': 0.8, 'pdf.fonttype': 42, 'ps.fonttype': 42})
    vals = corr.values
    labels = [str(c).replace('_', ' ').title() for c in corr.columns]
    n = len(labels)
    fig, ax = plt.subplots(figsize=(4.9, 4.4), dpi=300)
    fig.subplots_adjust(left=0.26, bottom=0.26, right=0.86, top=0.90 if title else 0.97)
    cmap = plt.get_cmap('RdBu_r')
    im = ax.imshow(vals, cmap=cmap, vmin=-1, vmax=1, interpolation='nearest', aspect='equal')
    for k in range(1, n):
        ax.axhline(k - 0.5, color='white', lw=0.5)
        ax.axvline(k - 0.5, color='white', lw=0.5)
    cbar = fig.colorbar(im, ax=ax, fraction=0.046, pad=0.04, shrink=0.85)
    cbar.set_label('Pearson Correlation', fontsize=7, labelpad=4)
    cbar.ax.tick_params(labelsize=6, length=2, width=0.5)
    cbar.outline.set_linewidth(0.5)
    cbar.set_ticks([-1.0, -0.5, 0.0, 0.5, 1.0])
    norm = mcolors.Normalize(vmin=-1, vmax=1)
    for i in range(n):
        for j in range(n):
            v = vals[i, j]
            rgba = cmap(norm(v))
            lum = 0.2126 * rgba[0] + 0.7152 * rgba[1] + 0.0722 * rgba[2]
            ax.text(j, i, f'{v:.2f}', ha='center', va='center', fontsize=6.5,
                    fontweight='bold', color='white' if lum < 0.52 else 'black')
    ax.set_xticks(np.arange(n))
    ax.set_yticks(np.arange(n))
    ax.set_xticklabels(labels, rotation=45, ha='right', va='top', rotation_mode='anchor', fontsize=9)
    ax.set_yticklabels(labels, rotation=0, ha='right', va='center', fontsize=9)
    ax.tick_params(axis='both', which='both', length=0, pad=3)
    for sp in ax.spines.values():
        sp.set_linewidth(0.8)
    if title:
        ax.set_title(title, fontsize=9, pad=6)
    fig.savefig(out_png, dpi=300, bbox_inches='tight', pad_inches=0.05, facecolor='white')
    plt.close(fig)


# Not exposed in the GUI. Kept as a library function so the Figure 3 reproduction
# scripts can still import it.
def export_merged_correlation_table(sessions, organ, behavior, bilateral, out_xlsx):
    '''Export the Figure-3 correlation INPUT as a reviewer-readable merged table.

    For each session's <organ>_summary.xlsx, keep the frames where the pain behavior
    occurs (`behavior`: ear='PB', eye='OT', nose='Bul'), RAW-merge the two sides per
    parameter (both->mean, one->that side; nose is single-sided) WITHOUT the mirror
    filter, and pool all sessions. This is exactly what the Figure-3 heatmap is built
    from — so a reviewer can read every number and recompute the matrix.

    sessions      : list of (name, folder) pairs.
    organ         : 'ear' | 'eye' | 'nose'.
    behavior      : the pain-behavior event label to keep.
    bilateral     : True for ear/eye (two sides), False for nose.
    out_xlsx      : output .xlsx path.

    Sheets written:
      - 'Merged_per_frame'   : Recording, Time, EventSide + 9 merged params (the pooled rows).
      - 'Per_recording_mean' : per session n_frames + mean of each merged param.
      - 'Correlation_9x9'    : Pearson matrix recomputed from Merged_per_frame (== heatmap).
    Also writes '<out_xlsx stem>_heatmap.png' (Figure-3 style) next to the workbook.
    Returns the pooled DataFrame (or None if nothing matched).'''
    fname = organ + '_summary.xlsx'
    evt = {'ear': 'Ear{s}Event', 'eye': 'Eye{s}Event', 'nose': 'NoseEvent'}[organ]
    frames = []
    for name, folder in sessions:
        fp = os.path.join(folder, fname)
        if not os.path.isfile(fp):
            continue
        df = pd.read_excel(fp)
        if bilateral:
            ev0, ev1 = evt.format(s=0), evt.format(s=1)
            if ev0 not in df.columns:
                continue
            m0 = (df[ev0] == behavior)
            m1 = (df[ev1] == behavior)
            sub = df[m0 | m1]
            if not len(sub):
                continue
            side = np.where(m0[m0 | m1] & m1[m0 | m1], 'both',
                            np.where(m0[m0 | m1], 'side0', 'side1'))
            out = {'Recording': name,
                   'Time': sub['Time'].values if 'Time' in sub.columns else np.arange(len(sub)),
                   'EventSide': side}
            for p in _CORR_PARAMS:
                d0 = pd.to_numeric(sub[_corr_col(p) + ' 0'], errors='coerce')
                d1 = pd.to_numeric(sub[_corr_col(p) + ' 1'], errors='coerce')
                s = d0.copy()
                both = d0.notna() & d1.notna()
                s[both] = (d0[both] + d1[both]) / 2
                only1 = d0.isna() & d1.notna()
                s[only1] = d1[only1]
                out[p] = s.values
            frames.append(pd.DataFrame(out))
        else:
            if 'NoseEvent' not in df.columns:
                continue
            sub = df[df['NoseEvent'] == behavior]
            if not len(sub):
                continue
            out = {'Recording': name,
                   'Time': sub['Time'].values if 'Time' in sub.columns else np.arange(len(sub)),
                   'EventSide': 'single'}
            for p in _CORR_PARAMS:
                out[p] = pd.to_numeric(sub[_corr_col(p)], errors='coerce').values
            frames.append(pd.DataFrame(out))
    if not frames:
        return None
    pooled = pd.concat(frames, ignore_index=True)
    corr = pooled[_CORR_PARAMS].corr(method='pearson')
    per_rec = (pooled.groupby('Recording')[_CORR_PARAMS].agg(['count', 'mean'])
               if len(pooled) else None)

    beh_name = {'PB': 'pulling behind (ear)', 'OT': 'orbital tightening (eye)',
                'Bul': 'bulging (nose)'}.get(behavior, behavior)
    about = pd.DataFrame({'Field': [
        'What this file is',
        'PAIN-related?',
        'Rows kept',
        'Merge rule',
        'Mirror filter',
        'Prerequisite',
        'Organ', 'Pain behavior', 'Recordings pooled', 'Frames pooled',
        'Correlation sheet'],
        'Value': [
        'Pearson correlation of the 9 facial parameters, as used in Figure 3.',
        'YES. Every row is a frame DURING the pain behavior, so this correlation '
        'describes how the facial parameters co-vary while the animal is in pain.',
        f"Only frames where the pain behavior '{behavior}' ({beh_name}) occurs "
        f"({'either side' if bilateral else 'nose'}).",
        'Two sides both present -> mean; only one present -> that side; '
        + ('nose is single-sided.' if not bilateral else 'none -> excluded.'),
        'NOT applied here (raw merge) — this is the correlation input, which differs '
        'from the pain-score input (mirror-filtered, all frames, intensity_area only).',
        'Built from the Step-1 summary files (ear/eye/nose_summary.xlsx). '
        'You must generate those first (Step 1) before this analysis.',
        organ, f'{behavior} ({beh_name})',
        int(pooled['Recording'].nunique()), int(len(pooled)),
        "'Correlation_9x9' = Pearson matrix recomputed from 'Merged_per_frame'."]})

    stem = os.path.splitext(out_xlsx)[0]
    os.makedirs(os.path.dirname(out_xlsx) or '.', exist_ok=True)
    # Write to a temp file in the same folder (valid .xlsx extension so pandas accepts
    # it), then atomically replace — so if the target is currently open in a viewer
    # (locked), we never truncate it on failure.
    tmp_xlsx = stem + '.writing.xlsx'
    with pd.ExcelWriter(tmp_xlsx, engine='openpyxl') as w:
        about.to_excel(w, sheet_name='About', index=False)
        corr.round(4).to_excel(w, sheet_name='Correlation_9x9')
        if per_rec is not None:
            flat = per_rec.copy()
            flat.columns = [f'{p}_{stat}' for p, stat in flat.columns]
            flat.reset_index().round(4).to_excel(w, sheet_name='Per_recording_mean', index=False)
        pooled.round(4).to_excel(w, sheet_name='Merged_per_frame', index=False)
    os.replace(tmp_xlsx, out_xlsx)

    # Figure-3 style heatmap PNG next to the workbook (temp + replace, same reason).
    try:
        png = stem + '_heatmap.png'
        tmp_png = stem + '_heatmap.writing.png'
        _correlation_heatmap(corr, tmp_png,
                             title=f'{organ.title()} facial parameters during {beh_name}')
        os.replace(tmp_png, png)
    except Exception:
        pass
    return pooled


# ============================================================================
# Video overlay writer
# ============================================================================

def write_overlay_video(video_path, per_frame_scores, output_path, score_times=None):
    '''
    Read *video_path* frame by frame, draw the pain score in the top-right
    corner for every frame whose score is not NaN, and write the result to
    *output_path* (.mp4).

    score_times : the real time (seconds) of each entry in per_frame_scores,
        i.e. data['time']. When given, each video frame is matched to the score
        whose time is closest to that frame's own timestamp (frame_index / fps).
        This keeps the burned-in number aligned with the animal's facial state
        even when the annotated video and the analysis ran at different frame
        rates -- e.g. the annotated .avi is 18000 frames at 30 fps while the
        analysis produced 17400 scores at 29 fps over the same 600 s. Matching by
        index instead would let the two drift apart by ~1 s per 30 s. Pass it
        whenever it is available; omit it only to reproduce the old index-based
        behaviour.

    Returns (frames_written, warning_message).  warning_message is '' on
    success, or a human-readable string when the frame counts differ and no
    time axis was supplied to reconcile them.
    '''
    import cv2

    cap = cv2.VideoCapture(video_path)
    if not cap.isOpened():
        raise IOError(f'Cannot open video: {video_path}')

    fps    = cap.get(cv2.CAP_PROP_FPS) or 30.0
    width  = int(cap.get(cv2.CAP_PROP_FRAME_WIDTH))
    height = int(cap.get(cv2.CAP_PROP_FRAME_HEIGHT))
    v_frames = int(cap.get(cv2.CAP_PROP_FRAME_COUNT))
    s_frames = len(per_frame_scores)

    times = None if score_times is None else np.asarray(score_times, dtype=float)

    warning = ''
    if times is None and v_frames != s_frames:
        warning = (
            f'Frame count mismatch: video has {v_frames} frames, '
            f'summary has {s_frames} frames, and no time axis was given to align '
            f'them. Overlay is applied by frame index up to frame '
            f'{min(v_frames, s_frames)}, which drifts when the two frame rates '
            f'differ. Pass score_times to align by time.'
        )

    fourcc = cv2.VideoWriter_fourcc(*'mp4v')
    out    = cv2.VideoWriter(output_path, fourcc, fps, (width, height))
    if not out.isOpened():
        # A freshly forked worker can hand back a VideoWriter that never opened
        # (the FFmpeg backend is not warmed up yet); it then silently writes
        # nothing. Retry once so the caller gets a real file or a clear error
        # instead of an empty output.
        out = cv2.VideoWriter(output_path, fourcc, fps, (width, height))
        if not out.isOpened():
            cap.release()
            raise IOError(f'Cannot open video writer for: {output_path}')

    font       = cv2.FONT_HERSHEY_SIMPLEX
    font_scale = 2.6
    thickness  = 4
    margin     = 12

    def score_for(frame_index):
        '''The score to draw on this video frame.'''
        if times is None:
            return per_frame_scores[frame_index] if frame_index < s_frames else np.nan
        # Nearest score by time: this frame sits at frame_index / fps seconds.
        t = frame_index / fps
        j = int(np.searchsorted(times, t))
        if j > 0 and (j >= len(times) or t - times[j - 1] <= times[j] - t):
            j -= 1
        return per_frame_scores[j]

    written = 0
    i = 0
    while True:
        ret, frame = cap.read()
        if not ret:
            break
        s = score_for(i)
        if not np.isnan(s):
            text = f'Pain: {s:.1f}'
            (tw, th), _ = cv2.getTextSize(text, font, font_scale, thickness)
            x = width  - tw - margin
            y = th + margin
            # Black outline (draw 4 offsets)
            for dx, dy in ((-1, -1), (-1, 1), (1, -1), (1, 1)):
                cv2.putText(frame, text, (x + dx, y + dy),
                            font, font_scale, (0, 0, 0), thickness, cv2.LINE_AA)
            # Bright red fill
            cv2.putText(frame, text, (x, y),
                        font, font_scale, (0, 0, 255), thickness, cv2.LINE_AA)
        out.write(frame)
        written += 1
        i += 1

    cap.release()
    out.release()
    return written, warning


# ============================================================================
# Animal Naming Dialog
# ============================================================================

class AnimalNamingDialog(wx.Dialog):
    '''
    Shows a scrollable list of detected animal subfolders.
    User types a name for each one. Default = subfolder name.
    '''

    def __init__(self, parent, detected):
        '''detected: [(default_name, folder_path), ...]'''
        super().__init__(parent, title='Name Each Animal', size=(720, 420),
                         style=wx.DEFAULT_DIALOG_STYLE | wx.RESIZE_BORDER)
        self.detected = detected
        self.name_inputs = []
        self._build_ui()

    def _build_ui(self):
        root = wx.BoxSizer(wx.VERTICAL)

        header = wx.StaticText(
            self,
            label=f'{len(self.detected)} animal folder(s) detected.\n'
                  'Edit the name for each animal below (default = folder name).',
        )
        root.Add(header, 0, wx.ALL, 12)

        scroll = scrolled.ScrolledPanel(self, size=(-1, 280))
        scroll.SetupScrolling()
        grid = wx.FlexGridSizer(cols=2, hgap=10, vgap=8)
        grid.AddGrowableCol(1, 1)

        for default_name, folder_path in self.detected:
            lbl = wx.StaticText(scroll, label=default_name, size=(280, -1))
            lbl.SetToolTip(folder_path)
            txt = wx.TextCtrl(scroll, value=default_name, size=(340, -1))
            grid.Add(lbl, 0, wx.ALIGN_CENTER_VERTICAL)
            grid.Add(txt, 1, wx.EXPAND)
            self.name_inputs.append(txt)

        scroll.SetSizer(grid)
        root.Add(scroll, 1, wx.LEFT | wx.RIGHT | wx.EXPAND, 12)
        root.Add(0, 8, 0)

        btn_row = wx.StdDialogButtonSizer()
        btn_ok     = wx.Button(self, wx.ID_OK,     label='Confirm')
        btn_cancel = wx.Button(self, wx.ID_CANCEL, label='Cancel')
        btn_row.AddButton(btn_ok)
        btn_row.AddButton(btn_cancel)
        btn_row.Realize()
        root.Add(btn_row, 0, wx.ALL | wx.ALIGN_RIGHT, 10)

        self.SetSizer(root)
        self.Centre()

    def get_names(self):
        return [(self.name_inputs[i].GetValue().strip() or self.detected[i][0],
                 self.detected[i][1])
                for i in range(len(self.detected))]


# ============================================================================
# Main Window
# ============================================================================

class InitialWindow(wx.Frame):

    def __init__(self, title):
        super(InitialWindow, self).__init__(parent=None, title=title, size=(750, 440))
        self.display_window()

    def display_window(self):
        panel = wx.Panel(self)
        boxsizer = wx.BoxSizer(wx.VERTICAL)

        self.text_welcome = wx.StaticText(panel, label='Welcome to LabGrymace!',
                                          style=wx.ALIGN_CENTER | wx.ST_ELLIPSIZE_END)
        boxsizer.Add(0, 60, 0)
        boxsizer.Add(self.text_welcome, 0, wx.LEFT | wx.RIGHT | wx.EXPAND, 5)
        boxsizer.Add(0, 60, 0)

        self.text_developers = wx.StaticText(
            panel,
            label='Developed by Wenjin Dong\n\nBing Ye Lab, Life Sciences Institute, University of Michigan',
            style=wx.ALIGN_CENTER | wx.ST_ELLIPSIZE_END,
        )
        boxsizer.Add(self.text_developers, 0, wx.LEFT | wx.RIGHT | wx.EXPAND, 5)
        boxsizer.Add(0, 60, 0)

        links = wx.BoxSizer(wx.HORIZONTAL)
        homepage  = hl.HyperLinkCtrl(panel, 0, 'Home Page',
                                     URL='https://github.com/devindwj0304/LabGym-LabGrymace')
        links.Add(homepage,  0, wx.LEFT | wx.RIGHT | wx.EXPAND, 10)
        boxsizer.Add(links, 0, wx.ALIGN_CENTER, 50)
        boxsizer.Add(0, 50, 0)

        button_facial = wx.Button(panel, label='Select Facial Expression Methods', size=(260, 40))
        button_facial.Bind(wx.EVT_BUTTON, self.window_facial)
        wx.Button.SetToolTip(button_facial,
                             'Select a facial expression analysis method to quantify pain scores.')
        boxsizer.Add(button_facial, 0, wx.ALIGN_CENTER)
        boxsizer.Add(0, 50, 0)

        panel.SetSizer(boxsizer)
        self.Centre()
        self.Show(True)

    def window_facial(self, event):
        WindowLv1_FacialMethods('Select Facial Expression Methods')


# ============================================================================
# Level-1: method selector
# ============================================================================

class WindowLv1_FacialMethods(wx.Frame):

    def __init__(self, title):
        super(WindowLv1_FacialMethods, self).__init__(parent=None, title=title, size=(500, 340))
        self.display_window()

    def display_window(self):
        panel = wx.Panel(self)
        boxsizer = wx.BoxSizer(wx.VERTICAL)
        boxsizer.Add(0, 30, 0)

        label = wx.StaticText(panel, label='Select a task:', style=wx.ALIGN_CENTER)
        boxsizer.Add(label, 0, wx.ALIGN_CENTER)
        boxsizer.Add(0, 18, 0)

        btn_gen = wx.Button(panel, label='Step 1 — Generate Summary Files', size=(320, 40))
        btn_gen.Bind(wx.EVT_BUTTON, self.open_generate)
        wx.Button.SetToolTip(
            btn_gen,
            'Select a folder of LabGym _processed output (raw behavior subfolders).\n'
            'LabGrymace will generate ear_summary.xlsx / eye_summary.xlsx / nose_summary.xlsx\n'
            'inside each animal folder.  Run this once per dataset.',
        )
        boxsizer.Add(btn_gen, 0, wx.ALIGN_CENTER)
        boxsizer.Add(0, 14, 0)

        button_mgs = wx.Button(panel, label='Step 2 — Pain Score  (Mouse Grimace Scale)', size=(320, 40))
        button_mgs.Bind(wx.EVT_BUTTON, self.open_pain_score)
        wx.Button.SetToolTip(
            button_mgs,
            'Compute pain scores from pre-generated summary files '
            '(ear_summary.xlsx, eye_summary.xlsx, nose_summary.xlsx).\n'
            'Uses CNO-calibrated Weighted Z-score Sum '
            f'(W_ear={W_EAR:.3f}, W_eye={W_EYE:.3f}, W_nose={W_NOSE:.3f}).',
        )
        boxsizer.Add(button_mgs, 0, wx.ALIGN_CENTER)
        boxsizer.Add(0, 30, 0)

        panel.SetSizer(boxsizer)
        self.Centre()
        self.Show(True)

    def open_generate(self, event):
        WindowLv2_GenerateSummary('Generate Summary Files')

    def open_pain_score(self, event):
        WindowLv2_PainScore('Pain Score')


# ============================================================================
# Level-2: Generate Summary Files window
# ============================================================================

class _GenerateSummaryDropTarget(wx.FileDropTarget):
    '''Accepts folders dragged from Finder onto the Generate Summary animal list.'''

    def __init__(self, window):
        super().__init__()
        self.window = window

    def OnDropFiles(self, x, y, filenames):
        existing_paths = {p for _, p in self.window._detected}
        added = 0
        for path in filenames:
            if not os.path.isdir(path):
                continue
            new_animals = find_raw_folders(path)
            if not new_animals and _is_raw_labgym_folder(Path(path)):
                new_animals = [(Path(path).name, path)]
            for name, fpath in new_animals:
                if fpath not in existing_paths:
                    self.window._detected.append((name, fpath))
                    self.window.list_animals.Append(name)
                    existing_paths.add(fpath)
                    added += 1
        if added:
            wx.MessageBox(f'Added {added} folder(s) via drag-and-drop.',
                          'Folders Added', wx.OK | wx.ICON_INFORMATION, self.window)
        return True


class _PainScoreDropTarget(wx.FileDropTarget):
    '''Accepts folders dragged from Finder onto the Pain Score animal list.'''

    def __init__(self, window):
        super().__init__()
        self.window = window

    def OnDropFiles(self, x, y, filenames):
        all_detected = []
        seen = set()
        for path in filenames:
            if not os.path.isdir(path):
                continue
            for name, fpath in find_animals_in_folder(path):
                if fpath not in seen:
                    all_detected.append((name, fpath))
                    seen.add(fpath)
        if not all_detected:
            wx.MessageBox(
                'No animal folders with summary files found.\n\n'
                'Each folder must contain:\n'
                '  ear_summary.xlsx, eye_summary.xlsx, nose_summary.xlsx\n\n'
                'Use Step 1 (Generate Summary Files) first.',
                'Nothing Found', wx.OK | wx.ICON_WARNING, self.window,
            )
            return True
        loaded = self.window._add_animals_from_detected(all_detected)
        if loaded:
            wx.MessageBox(f'Added {loaded} animal(s).',
                          'Done', wx.OK | wx.ICON_INFORMATION, self.window)
        return True


class WindowLv2_GenerateSummary(wx.Frame):
    '''
    Lets the user pick a folder of LabGym _processed output, then runs
    FacialDataIntegrator + LoadDataAdvanced to produce ear/eye/nose_summary.xlsx
    inside each animal subfolder.
    '''

    def __init__(self, title):
        super(WindowLv2_GenerateSummary, self).__init__(parent=None, title=title, size=(900, 680))
        self._detected   = []   # [(name, raw_folder_path), ...]
        self._output_dir = None
        self.display_window()

    def display_window(self):
        panel = wx.Panel(self)
        vbox = wx.BoxSizer(wx.VERTICAL)
        vbox.Add(0, 10, 0)

        # ── Row 1: add / remove raw data folders ──
        row1 = wx.BoxSizer(wx.HORIZONTAL)
        btn_col = wx.BoxSizer(wx.VERTICAL)
        btn_pick = wx.Button(panel, label='Add raw data folder', size=(220, 36))
        btn_pick.Bind(wx.EVT_BUTTON, self.pick_folder)
        wx.Button.SetToolTip(
            btn_pick,
            'Select a folder containing LabGym _processed animal subfolders.\n'
            'Each subfolder must contain behavior subfolders: fEarBL, fEyeOT, noseBul, etc.\n'
            'Click multiple times to add animals from different locations.',
        )
        btn_remove = wx.Button(panel, label='Remove selected', size=(220, 28))
        btn_remove.Bind(wx.EVT_BUTTON, self.remove_animal)
        btn_col.Add(btn_pick,   0, wx.BOTTOM, 4)
        btn_col.Add(btn_remove, 0)
        row1.Add(btn_col, 0, wx.LEFT | wx.RIGHT, 10)

        self.list_animals = wx.ListBox(panel, style=wx.LB_EXTENDED, size=(-1, 90))
        wx.ListBox.SetToolTip(self.list_animals,
                              'Detected animal folders. Drag folders onto this window from Finder, or click "Add raw data folder".')
        row1.Add(self.list_animals, 1, wx.LEFT | wx.RIGHT | wx.EXPAND, 5)
        vbox.Add(row1, 0, wx.LEFT | wx.RIGHT | wx.EXPAND, 10)
        vbox.Add(0, 8, 0)

        # ── Row 2: output folder picker ──
        row2 = wx.BoxSizer(wx.HORIZONTAL)
        btn_out = wx.Button(panel, label='Select output folder\n(for summary files)', size=(220, 50))
        btn_out.Bind(wx.EVT_BUTTON, self.pick_output)
        wx.Button.SetToolTip(
            btn_out,
            'Choose where the summary files will be saved.\n'
            'One subfolder per animal will be created here:\n'
            '  <output>/<animal_name>/ear_summary.xlsx\n'
            '  <output>/<animal_name>/eye_summary.xlsx\n'
            '  <output>/<animal_name>/nose_summary.xlsx\n\n'
            'In Step 2 (Pain Score), point to this output folder.',
        )
        self.lbl_output = wx.StaticText(panel, label='None  (will save alongside raw data).',
                                        style=wx.ALIGN_LEFT | wx.ST_ELLIPSIZE_END)
        row2.Add(btn_out,         0, wx.LEFT | wx.RIGHT, 10)
        row2.Add(self.lbl_output, 1, wx.LEFT | wx.RIGHT | wx.ALIGN_CENTER_VERTICAL, 5)
        vbox.Add(row2, 0, wx.LEFT | wx.RIGHT | wx.EXPAND, 10)
        vbox.Add(0, 10, 0)

        # ── Row 3: Generate button ──
        btn_gen = wx.Button(panel, label='Generate summary files for all listed animals', size=(380, 36))
        btn_gen.Bind(wx.EVT_BUTTON, self.run_generate)
        vbox.Add(btn_gen, 0, wx.LEFT, 20)
        vbox.Add(0, 8, 0)

        # ── Row 4: Log ──
        vbox.Add(wx.StaticText(panel, label='Log:'), 0, wx.LEFT, 20)
        vbox.Add(0, 4, 0)
        self.log = wx.TextCtrl(panel, style=wx.TE_MULTILINE | wx.TE_READONLY | wx.HSCROLL,
                               size=(-1, 150))
        vbox.Add(self.log, 1, wx.LEFT | wx.RIGHT | wx.BOTTOM | wx.EXPAND, 20)

        panel.SetSizer(vbox)
        self.Centre()
        self.SetDropTarget(_GenerateSummaryDropTarget(self))
        self.Show(True)

    # ── handlers ──

    def _log(self, msg):
        self.log.AppendText(msg + '\n')
        wx.GetApp().Yield()

    def pick_folder(self, event):
        dlg = wx.DirDialog(self, 'Select folder containing LabGym processed animal data',
                           style=wx.DD_DEFAULT_STYLE)
        if dlg.ShowModal() != wx.ID_OK:
            dlg.Destroy()
            return
        root = dlg.GetPath()
        dlg.Destroy()

        new_animals = find_raw_folders(root)
        if not new_animals:
            wx.MessageBox(
                'No raw LabGym animal folders found.\n\n'
                'Each subfolder must contain behavior subfolders such as:\n'
                '  fEarBL, sEarBL, fEyeOT, sEyeOT, noseBul, noseBL',
                'No Data Found', wx.OK | wx.ICON_WARNING,
            )
            return

        existing_paths = {p for _, p in self._detected}
        added = 0
        for name, path in new_animals:
            if path not in existing_paths:
                self._detected.append((name, path))
                self.list_animals.Append(name)
                existing_paths.add(path)
                added += 1
        if added == 0:
            wx.MessageBox('All detected animals from this folder are already in the list.',
                          'Already Added', wx.OK | wx.ICON_INFORMATION)

    def remove_animal(self, event):
        selections = list(self.list_animals.GetSelections())
        for idx in sorted(selections, reverse=True):
            self.list_animals.Delete(idx)
            del self._detected[idx]

    def pick_output(self, event):
        dlg = wx.DirDialog(self, 'Select folder to save summary files',
                           style=wx.DD_DEFAULT_STYLE | wx.DD_NEW_DIR_BUTTON)
        if dlg.ShowModal() == wx.ID_OK:
            self._output_dir = dlg.GetPath()
            self.lbl_output.SetLabel(self._output_dir)
        dlg.Destroy()

    def run_generate(self, event):
        if not self._detected:
            wx.MessageBox('Please select a raw data folder first.', 'Nothing to do',
                          wx.OK | wx.ICON_WARNING)
            return

        self.log.Clear()
        ok, skip, fail = 0, 0, 0
        available = []   # (name, out_dir) folders that now have summary files

        for name, folder_path in self._detected:
            # Determine output directory for this animal
            if self._output_dir:
                out_dir = str(Path(self._output_dir) / name)
            else:
                out_dir = folder_path   # save alongside raw data

            out = Path(out_dir)

            # Skip if already done
            if _has_summary_files(out):
                self._log(f'[SKIP]  {name}  — summary files already exist in {out_dir}')
                skip += 1
                available.append((name, out_dir))
                continue

            self._log(f'[GEN]   {name}  …')
            try:
                generate_summary_files(folder_path, output_dir=out_dir)
                if _has_summary_files(out):
                    self._log(f'        → saved to {out_dir}')
                    ok += 1
                    available.append((name, out_dir))
                else:
                    self._log(f'        → WARNING: generation ran but output files not found')
                    fail += 1
            except Exception as exc:
                import traceback
                self._log(f'        → ERROR: {exc}')
                self._log(traceback.format_exc())
                fail += 1

        self._log('')
        self._log(f'Done — {ok} generated, {skip} skipped, {fail} failed.')

        if fail == 0:
            dest = self._output_dir or '(each animal folder)'
            wx.MessageBox(
                f'Summary files generated successfully.\n\n'
                f'{ok} animal(s) processed, {skip} already existed.\n\n'
                f'Output location: {dest}\n'
                f'Now open Step 2 (Pain Score) and select that output folder.',
                'Done', wx.OK | wx.ICON_INFORMATION,
            )

class VideoAnimalPairDialog(wx.Dialog):
    '''
    Lets the user pick one video from the loaded video list and one animal
    from the loaded animal records, then click OK to proceed with overlay.
    '''

    def __init__(self, parent, video_paths, animal_records):
        super().__init__(parent, title='Select Video and Animal to Pair',
                         style=wx.DEFAULT_DIALOG_STYLE)
        self._video_paths    = video_paths
        self._animal_records = animal_records

        panel  = wx.Panel(self)
        vbox   = wx.BoxSizer(wx.VERTICAL)
        vbox.Add(0, 12, 0)

        # Video choice
        row_v = wx.BoxSizer(wx.HORIZONTAL)
        row_v.Add(wx.StaticText(panel, label='Select video:'),
                  0, wx.LEFT | wx.ALIGN_CENTER_VERTICAL, 16)
        self.choice_video = wx.Choice(
            panel,
            choices=[os.path.basename(p) for p in video_paths],
            size=(340, -1),
        )
        self.choice_video.SetSelection(0)
        row_v.Add(self.choice_video, 1, wx.LEFT | wx.RIGHT, 10)
        vbox.Add(row_v, 0, wx.EXPAND | wx.BOTTOM, 10)

        # Animal choice
        row_a = wx.BoxSizer(wx.HORIZONTAL)
        row_a.Add(wx.StaticText(panel, label='Select animal:'),
                  0, wx.LEFT | wx.ALIGN_CENTER_VERTICAL, 16)
        self.choice_animal = wx.Choice(
            panel,
            choices=[r['name'] for r in animal_records],
            size=(340, -1),
        )
        self.choice_animal.SetSelection(0)
        row_a.Add(self.choice_animal, 1, wx.LEFT | wx.RIGHT, 10)
        vbox.Add(row_a, 0, wx.EXPAND | wx.BOTTOM, 16)

        # OK / Cancel
        btn_sizer = wx.StdDialogButtonSizer()
        btn_ok     = wx.Button(panel, wx.ID_OK)
        btn_cancel = wx.Button(panel, wx.ID_CANCEL)
        btn_sizer.AddButton(btn_ok)
        btn_sizer.AddButton(btn_cancel)
        btn_sizer.Realize()
        vbox.Add(btn_sizer, 0, wx.ALIGN_RIGHT | wx.RIGHT | wx.BOTTOM, 12)

        panel.SetSizer(vbox)
        vbox.Fit(self)
        self.Centre()

    def get_selection(self):
        vi = self.choice_video.GetSelection()
        ai = self.choice_animal.GetSelection()
        return self._video_paths[vi], self._animal_records[ai]


# ============================================================================
# Level-2: Pain Score window
# ============================================================================

class WindowLv2_PainScore(wx.Frame):

    def __init__(self, title):
        super(WindowLv2_PainScore, self).__init__(parent=None, title=title, size=(960, 720))
        self.output_path    = None
        self.animal_records = []   # list of dicts: name, folder, data, windows
        self._video_paths   = []   # list of selected LabGym analyzed video paths
        self.display_window()

    def display_window(self):
        panel = wx.Panel(self)
        boxsizer = wx.BoxSizer(wx.VERTICAL)
        boxsizer.Add(0, 10, 0)

        # ── Module A: Select LabGym Analyzed Videos ───────────────────────
        lbl_videos = wx.StaticText(panel, label='Select LabGym Analyzed Videos:')
        boxsizer.Add(lbl_videos, 0, wx.LEFT, 20)
        boxsizer.Add(0, 4, 0)

        module_videos = wx.BoxSizer(wx.HORIZONTAL)
        btn_vid_col = wx.BoxSizer(wx.VERTICAL)

        btn_add_videos = wx.Button(panel, label='Add LabGym analyzed videos', size=(300, 36))
        btn_add_videos.Bind(wx.EVT_BUTTON, self.add_videos)
        wx.Button.SetToolTip(
            btn_add_videos,
            'Select one or more video files (.avi, .mp4) generated by\n'
            'LabGym\'s "Analyze Videos" function.\n'
            'Click multiple times to add videos from different folders.',
        )
        btn_remove_videos = wx.Button(panel, label='Remove selected video', size=(300, 28))
        btn_remove_videos.Bind(wx.EVT_BUTTON, self.remove_selected_video)

        btn_vid_col.Add(btn_add_videos,    0, wx.BOTTOM, 4)
        btn_vid_col.Add(btn_remove_videos, 0)

        self.list_videos = wx.ListBox(panel, style=wx.LB_EXTENDED, size=(580, 90))
        wx.ListBox.SetToolTip(self.list_videos, 'Selected LabGym analyzed video files.')

        module_videos.Add(btn_vid_col,      0, wx.LEFT | wx.RIGHT, 10)
        module_videos.Add(self.list_videos, 1, wx.LEFT | wx.RIGHT | wx.EXPAND, 5)
        boxsizer.Add(module_videos, 0, wx.LEFT | wx.RIGHT | wx.EXPAND, 10)
        boxsizer.Add(0, 12, 0)

        # ── Module B: Add Folder with Summary Files ───────────────────────
        lbl_summary = wx.StaticText(panel, label='Add Folder with Summary Files:')
        boxsizer.Add(lbl_summary, 0, wx.LEFT, 20)
        boxsizer.Add(0, 4, 0)

        module_top = wx.BoxSizer(wx.HORIZONTAL)

        btn_col = wx.BoxSizer(wx.VERTICAL)
        button_add = wx.Button(
            panel,
            label='Add folder with summary files',
            size=(300, 36),
        )
        button_add.Bind(wx.EVT_BUTTON, self.add_folder)
        wx.Button.SetToolTip(
            button_add,
            'Select a folder whose subfolders each contain pre-generated summary files:\n'
            '  ear_summary.xlsx, eye_summary.xlsx, nose_summary.xlsx\n\n'
            'If you have not generated these yet, use Step 1 first.\n'
            'You can also select a single animal folder directly.\n'
            'Click multiple times to add animals from different locations.',
        )
        button_remove = wx.Button(panel, label='Remove selected animal', size=(300, 28))
        button_remove.Bind(wx.EVT_BUTTON, self.remove_selected)

        btn_col.Add(button_add,    0, wx.BOTTOM, 4)
        btn_col.Add(button_remove, 0)

        self.list_animals = wx.ListBox(panel, style=wx.LB_SINGLE, size=(580, 90))
        wx.ListBox.SetToolTip(self.list_animals,
                              'Loaded animals. Drag folders onto this window from Finder to add multiple at once.')

        module_top.Add(btn_col,           0, wx.LEFT | wx.RIGHT, 10)
        module_top.Add(self.list_animals, 1, wx.LEFT | wx.RIGHT | wx.EXPAND, 5)
        boxsizer.Add(module_top, 0, wx.LEFT | wx.RIGHT | wx.EXPAND, 10)
        boxsizer.Add(0, 10, 0)

        # ── Row 3b: Analysis time range (seconds) ─────────────────────────
        module_range = wx.BoxSizer(wx.HORIZONTAL)
        self.chk_range = wx.CheckBox(panel, label='Restrict analysis to time range (s):',
                                     size=(300, -1))
        self.chk_range.SetValue(True)          # default = 0-500 s (the Figure 6 window)
        self.spin_start = wx.SpinCtrl(panel, min=0, max=100000, initial=0,   size=(90, -1))
        lbl_to = wx.StaticText(panel, label='to')
        self.spin_end   = wx.SpinCtrl(panel, min=1, max=100000, initial=500, size=(90, -1))
        sec_per_win = FRAME_WINDOW / 30.0
        for w in (self.chk_range, self.spin_start, self.spin_end):
            wx.Window.SetToolTip(
                w,
                f'Restrict every output (per-window sheet, per-frame sheet, Summary and\n'
                f'the plot) to this time range. Enter any start/end seconds you like; the range\n'
                f'snaps to whole {sec_per_win:.0f} s windows. Defaults to 0-500 s (windows 1-5).\n'
                f'Uncheck for the full recording.',
            )
        module_range.Add(self.chk_range,  0, wx.LEFT | wx.RIGHT | wx.ALIGN_CENTER_VERTICAL, 10)
        module_range.Add(self.spin_start, 0, wx.RIGHT | wx.ALIGN_CENTER_VERTICAL, 5)
        module_range.Add(lbl_to,          0, wx.RIGHT | wx.ALIGN_CENTER_VERTICAL, 5)
        module_range.Add(self.spin_end,   0, wx.RIGHT | wx.ALIGN_CENTER_VERTICAL, 5)
        boxsizer.Add(module_range, 0, wx.LEFT | wx.RIGHT | wx.EXPAND, 10)
        boxsizer.Add(0, 10, 0)

        # ── Row 3c: Mirror-reflection filter (opt-in) ─────────────────────
        module_mirror = wx.BoxSizer(wx.HORIZONTAL)
        self.chk_mirror = wx.CheckBox(
            panel, label='Apply mirror-reflection filter (opt-in)', size=(300, -1))
        self.chk_mirror.SetValue(APPLY_MIRROR_FILTER)   # default = OFF
        wx.Window.SetToolTip(
            self.chk_mirror,
            'Leave OFF for normal recordings. Check this ONLY when the LabGym\n'
            'tracker may have latched onto mirror reflections (spurious high-velocity\n'
            'jumps toward the mirror). When ON, frames flagged by the 4-condition\n'
            'mirror filter are removed before the pain score is computed and are\n'
            'highlighted orange in the Merge-signals outputs.',
        )
        module_mirror.Add(self.chk_mirror, 0,
                          wx.LEFT | wx.RIGHT | wx.ALIGN_CENTER_VERTICAL, 10)
        boxsizer.Add(module_mirror, 0, wx.LEFT | wx.RIGHT | wx.EXPAND, 10)
        boxsizer.Add(0, 10, 0)

        # ── Row 4: Output folder ──────────────────────────────────────────
        module_out = wx.BoxSizer(wx.HORIZONTAL)
        button_out = wx.Button(panel, label='Select a folder to store\nthe results', size=(300, 40))
        button_out.Bind(wx.EVT_BUTTON, self.select_output)
        wx.Button.SetToolTip(
            button_out,
            'Outputs: pain_scores.xlsx (per-animal sheets + summary), '
            'pain_scores_per_frame.xlsx, and pain_score_chart.png.',
        )
        self.text_output = wx.StaticText(panel, label='None.',
                                         style=wx.ALIGN_LEFT | wx.ST_ELLIPSIZE_END)
        module_out.Add(button_out,       0, wx.LEFT | wx.RIGHT, 10)
        module_out.Add(self.text_output, 1, wx.LEFT | wx.RIGHT | wx.ALIGN_CENTER_VERTICAL, 5)
        boxsizer.Add(module_out, 0, wx.LEFT | wx.RIGHT | wx.EXPAND, 10)
        boxsizer.Add(0, 25, 0)

        # ── Start button ───────────────────────────────────────────────────
        button_start = wx.Button(panel, label='Calculate pain score and generate chart', size=(330, 40))
        button_start.Bind(wx.EVT_BUTTON, self.calculate_and_plot)
        wx.Button.SetToolTip(
            button_start,
            f'Divide each recording into {FRAME_WINDOW}-frame windows. '
            'Compute mean Intensity Area per window, then apply CNO-calibrated '
            'Z-scores with fixed normalization (0 = baseline, 100 = 1mg CNO). '
            'Values outside 0-100 are meaningful.',
        )
        boxsizer.Add(button_start, 0, wx.RIGHT | wx.ALIGN_RIGHT, 90)
        boxsizer.Add(0, 10, 0)

        button_overlay = wx.Button(panel, label='Generate Video with Pain Score Overlay', size=(330, 40))
        button_overlay.Bind(wx.EVT_BUTTON, self.generate_overlay)
        wx.Button.SetToolTip(
            button_overlay,
            'Select one analyzed video (Module A) and one animal\'s summary data (Module B) '
            'to generate a new video with per-frame pain score overlaid in the top-right corner.\n'
            'Score is computed using a causal 60-frame (2 s @30fps) look-back window.\n'
            'The first 59 frames have no score and no overlay.',
        )
        boxsizer.Add(button_overlay, 0, wx.RIGHT | wx.ALIGN_RIGHT, 90)
        boxsizer.Add(0, 10, 0)

        panel.SetSizer(boxsizer)
        self.Centre()
        self.SetDropTarget(_PainScoreDropTarget(self))
        self.Show(True)

    # ── Handlers ──────────────────────────────────────────────────────────

    def add_videos(self, event):
        wildcard = 'Video files (*.avi;*.mp4;*.mov)|*.avi;*.mp4;*.mov|All files (*.*)|*.*'
        dlg = wx.FileDialog(self, 'Select LabGym analyzed video files', wildcard=wildcard,
                            style=wx.FD_OPEN | wx.FD_MULTIPLE)
        if dlg.ShowModal() != wx.ID_OK:
            dlg.Destroy()
            return
        paths = dlg.GetPaths()
        dlg.Destroy()

        existing = set(self._video_paths)
        for p in paths:
            if p not in existing:
                self._video_paths.append(p)
                self.list_videos.Append(os.path.basename(p))
                existing.add(p)

    def remove_selected_video(self, event):
        selections = list(self.list_videos.GetSelections())
        for idx in sorted(selections, reverse=True):
            self.list_videos.Delete(idx)
            del self._video_paths[idx]

    def _add_animals_from_detected(self, detected):
        '''Show naming dialog then load data for the given detected list.
        Returns number of animals successfully loaded.'''
        name_dlg = AnimalNamingDialog(self, detected)
        if name_dlg.ShowModal() != wx.ID_OK:
            name_dlg.Destroy()
            return 0
        named = name_dlg.get_names()
        name_dlg.Destroy()

        existing_names = {r['name'] for r in self.animal_records}
        loaded = 0
        for animal_name, folder_path in named:
            final_name = animal_name
            suffix = 1
            while final_name in existing_names:
                final_name = f'{animal_name}_{suffix}'
                suffix += 1

            data = load_raw_data(folder_path,
                                 apply_mirror_filter=self.chk_mirror.GetValue())
            if data is not None:
                print(f'[{animal_name}] filter: {data.get("filter_source", "unknown")}')

            if data is None:
                wx.MessageBox(
                    f'Error loading "{animal_name}".\n\n'
                    'Could not read ear_summary.xlsx / eye_summary.xlsx / nose_summary.xlsx.\n'
                    'Run Step 1 (Generate Summary Files) on this folder first.',
                    'Load Error', wx.OK | wx.ICON_ERROR)
                continue

            windows = compute_windowed_pain_scores(data)
            if not windows:
                wx.MessageBox(
                    f'"{animal_name}": recording too short.\n'
                    f'Need at least {FRAME_WINDOW} frames, got {data["n_frames"]}.',
                    'Load Warning', wx.OK | wx.ICON_WARNING,
                )
                continue

            self.animal_records.append({
                'name':    final_name,
                'folder':  folder_path,
                'data':    data,
                'windows': windows,
            })
            self.list_animals.Append(
                f'{final_name}   ({data["n_frames"]} frames  →  {len(windows)} windows)'
            )
            existing_names.add(final_name)
            loaded += 1

        return loaded

    def add_folder(self, event):
        dlg = wx.DirDialog(self, 'Select folder containing summary files',
                           style=wx.DD_DEFAULT_STYLE)
        if dlg.ShowModal() != wx.ID_OK:
            dlg.Destroy()
            return
        root = dlg.GetPath()
        dlg.Destroy()

        detected = find_animals_in_folder(root)
        if not detected:
            wx.MessageBox(
                'No animal folders with summary files found.\n\n'
                'Each subfolder must contain:\n'
                '  ear_summary.xlsx\n'
                '  eye_summary.xlsx\n'
                '  nose_summary.xlsx\n\n'
                'Use Step 1 (Generate Summary Files) to create them first.',
                'No Data Found', wx.OK | wx.ICON_WARNING,
            )
            return

        loaded = self._add_animals_from_detected(detected)
        if loaded:
            wx.MessageBox(f'Added {loaded} animal(s).', 'Done', wx.OK | wx.ICON_INFORMATION)

    def remove_selected(self, event):
        idx = self.list_animals.GetSelection()
        if idx == wx.NOT_FOUND:
            wx.MessageBox('Please select an animal to remove.', 'Nothing Selected',
                          wx.OK | wx.ICON_INFORMATION)
            return
        self.list_animals.Delete(idx)
        del self.animal_records[idx]

    def select_output(self, event):
        dlg = wx.DirDialog(self, 'Select output folder', style=wx.DD_DEFAULT_STYLE)
        if dlg.ShowModal() == wx.ID_OK:
            self.output_path = dlg.GetPath()
            self.text_output.SetLabel(self.output_path)
        dlg.Destroy()

    def calculate_and_plot(self, event):
        if not self.animal_records:
            wx.MessageBox('Please add at least one animal folder.', 'Missing Input',
                          wx.OK | wx.ICON_WARNING)
            return
        if not self.output_path:
            wx.MessageBox('Please select an output folder.', 'Missing Input',
                          wx.OK | wx.ICON_WARNING)
            return

        # Optional time-range restriction -> whole-window (w_start, w_end)
        window_range = None
        range_msg = 'full recording'
        if self.chk_range.GetValue():
            start_s = float(self.spin_start.GetValue())
            end_s   = float(self.spin_end.GetValue())
            if end_s <= start_s:
                wx.MessageBox('End time must be greater than start time.', 'Invalid Range',
                              wx.OK | wx.ICON_WARNING)
                return
            sec_per_win = FRAME_WINDOW / 30.0            # 100 s per window @30fps
            w_start = int(start_s // sec_per_win) + 1
            w_end   = int(np.ceil(end_s / sec_per_win))
            window_range = (w_start, w_end)
            snap_start = (w_start - 1) * sec_per_win     # window bounds actually analyzed
            snap_end   = w_end * sec_per_win
            range_msg = (f'{start_s:.0f}-{end_s:.0f} s  ->  windows {w_start}-{w_end} '
                         f'({snap_start:.0f}-{snap_end:.0f} s of data used)')
            # Non-zero start: the 2 s causal look-back reaches ~59 frames before the range,
            # so the first ~2 s of the range is not fully self-contained. (Figure 6 uses
            # 0-500 s / start=0, which is unaffected.) Warn and let the user decide.
            if w_start > 1:
                proceed = wx.MessageBox(
                    f'The range starts at {start_s:.0f} s (not 0). Pain score uses a causal '
                    f'2 s (60-frame) look-back, so the first ~2 s of the range still reflects '
                    f'a little data from just before {start_s:.0f} s.\n\n'
                    f'This is fine for a full 0 to N s window (the default is 0-500 s). '
                    f'Continue anyway?',
                    'Range does not start at 0 s', wx.OK | wx.CANCEL | wx.ICON_WARNING)
                if proceed != wx.OK:
                    return

        try:
            # All file writing is shared with the headless batch driver via
            # write_pain_score_outputs() so the GUI and scripted output are identical.
            write_pain_score_outputs(self.animal_records, self.output_path,
                                     window_range=window_range)
            n_animals = len(self.animal_records)

            wx.MessageBox(
                f'Done!  {n_animals} animal(s) processed.\n'
                f'Analysis range: {range_msg}\n\n'
                f'Spreadsheets:\n'
                f'  pain_scores.xlsx\n'
                f'  — one sheet per animal (window, time, intensities, Z-scores, pain_score)\n'
                f'  — "Summary" sheet (per-animal pain score over the selected range)\n'
                f'  pain_scores_per_frame.xlsx\n'
                f'  — one sheet per animal; per-frame pain score with its 2 s (60-frame)\n'
                f'    look-back window and full breakdown (win bounds, 2 s means, Z, raw, pain)\n\n'
                f'Plots:\n'
                f'  pain_score_chart.png  (one point per animal)\n\n'
                f'Saved to: {self.output_path}',
                'Success', wx.OK | wx.ICON_INFORMATION,
            )

        except Exception as e:
            import traceback
            wx.MessageBox(f'An error occurred:\n{e}\n\n{traceback.format_exc()}',
                          'Error', wx.OK | wx.ICON_ERROR)

    def generate_overlay(self, event):
        if not self._video_paths:
            wx.MessageBox('Please add at least one video in the "Select LabGym Analyzed Videos" section.',
                          'No Video', wx.OK | wx.ICON_WARNING)
            return
        if not self.animal_records:
            wx.MessageBox('Please add at least one animal folder in the "Add Folder with Summary Files" section.',
                          'No Animal', wx.OK | wx.ICON_WARNING)
            return
        if not self.output_path:
            wx.MessageBox('Please select an output folder first.',
                          'No Output Folder', wx.OK | wx.ICON_WARNING)
            return

        dlg = VideoAnimalPairDialog(self, self._video_paths, self.animal_records)
        if dlg.ShowModal() != wx.ID_OK:
            dlg.Destroy()
            return
        video_path, record = dlg.get_selection()
        dlg.Destroy()

        animal_name = record['name']
        out_path    = os.path.join(self.output_path, f'{animal_name}_pain_overlay.mp4')

        try:
            scores = compute_per_frame_pain_scores(record['data'])

            # Show a non-blocking busy message
            busy = wx.BusyInfo('Generating overlay video, please wait…', parent=self)
            wx.GetApp().Yield()

            written, warning = write_overlay_video(
                video_path, scores, out_path, score_times=record['data']['time'])
            del busy

            if warning:
                wx.MessageBox(warning, 'Warning', wx.OK | wx.ICON_WARNING)

            wx.MessageBox(
                f'Done!  {written} frames written.\n\n'
                f'Output: {out_path}',
                'Overlay Video Generated', wx.OK | wx.ICON_INFORMATION,
            )

        except Exception as e:
            import traceback
            wx.MessageBox(f'An error occurred:\n{e}\n\n{traceback.format_exc()}',
                          'Error', wx.OK | wx.ICON_ERROR)


# ============================================================================
# Entry point
# ============================================================================

def main_window():
    app = wx.App()
    InitialWindow(f'LabGrymace v{__version__}')
    app.MainLoop()


if __name__ == '__main__':
    main_window()
